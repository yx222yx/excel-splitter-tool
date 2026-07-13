from pathlib import Path
import sys
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "人员数据.xlsx"
    workbook = Workbook()

    first = workbook.active
    first.title = "人员"
    first.merge_cells("A1:D1")
    first["A1"] = "人员奖金明细"
    first["A1"].font = Font(bold=True, size=14)
    first["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    first.append(["部门", "姓名", "金额", "计算值"])
    first.append(["临床部", "张三", 100, "=C3*2"])
    first.append([" \u3000研发部 ", "李四", 200, "=C4*2"])
    first.append([None, "空值人员", 300, "=C5*2"])
    first.freeze_panes = "B3"
    first.column_dimensions["A"].width = 18
    first.row_dimensions[1].height = 24

    second = workbook.create_sheet("项目")
    second.append(["所属部门", "项目"])
    second.append(["临床部", "项目A"])
    second.append(["研发部", "项目B"])
    second.append(["市场部", "项目C"])

    notes = workbook.create_sheet("说明")
    notes["A1"] = "该 sheet 不参与拆分"

    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def formula_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "公式字段.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "公式数据"
    sheet.append(["部门公式", "姓名", "计算结果"])
    sheet.append(['=IF(1=1,"临床部","")', "张三", "=1+1"])
    sheet.append(['=IF(1=1,"研发部","")', "李四", "=2+2"])
    sheet.append([None, "空值人员", 0])
    workbook.save(path)
    workbook.close()

    _set_formula_cache(
        path,
        {"A2": "临床部", "A3": "研发部", "C2": 2, "C3": 4},
    )
    return path


@pytest.fixture
def late_header_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "第八行表头.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "晚表头"
    for index in range(1, 8):
        sheet.append([f"标题说明 {index}", None, None])
    sheet.append(["部门", "姓名", "金额"])
    sheet.append(["临床部", "张三", 100])
    sheet.append(["研发部", "李四", 200])
    sheet.append(["市场部", "王五", 300])
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def fourteenth_header_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "第十四行表头.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "更晚表头"
    for index in range(1, 14):
        sheet.append([f"标题说明 {index}", None, None])
    sheet.append(["部门", "姓名", "金额"])
    sheet.append(["临床部", "张三", 100])
    sheet.append(["研发部", "李四", 200])
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def linked_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "关联拆分.xlsx"
    workbook = Workbook()

    reference = workbook.active
    reference.title = "人员归属"
    reference.append(["二级团队", "姓名", "奖金"])
    reference.append(["团队甲", "张三", 100])
    reference.append(["团队甲", "李四", 200])
    reference.append(["团队乙", "王五", 300])

    detail = workbook.create_sheet("工时明细")
    detail.append(["员工姓名", "项目"])
    detail.append(["张三", "项目A"])
    detail.append(["李四", "项目B"])
    detail.append(["王五", "项目C"])
    detail.append(["未登记人员", "项目D"])
    detail.append([None, "项目E"])

    direct = workbook.create_sheet("团队汇总")
    direct.append(["二级团队", "指标"])
    direct.append(["团队甲", 10])
    direct.append(["团队乙", 20])

    notes = workbook.create_sheet("说明")
    notes.append(["说明"])
    notes.append(["完整保留第一行"])
    notes.append(["完整保留第二行"])

    workbook.save(path)
    workbook.close()
    return path


def _set_formula_cache(path: Path, cached_values: dict[str, object]) -> None:
    with ZipFile(path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}

    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root = ElementTree.fromstring(entries["xl/worksheets/sheet1.xml"])
    for cell in root.findall(f".//{{{namespace}}}c"):
        reference = cell.get("r")
        if reference not in cached_values:
            continue
        value = cached_values[reference]
        if isinstance(value, str):
            cell.set("t", "str")
        else:
            cell.attrib.pop("t", None)
        value_node = cell.find(f"{{{namespace}}}v")
        if value_node is None:
            value_node = ElementTree.SubElement(cell, f"{{{namespace}}}v")
        value_node.text = str(value)

    entries["xl/worksheets/sheet1.xml"] = ElementTree.tostring(
        root, encoding="utf-8", xml_declaration=True
    )
    temporary = path.with_suffix(".tmp")
    with ZipFile(temporary, "w", ZIP_DEFLATED) as target:
        for name, content in entries.items():
            target.writestr(name, content)
    temporary.replace(path)
