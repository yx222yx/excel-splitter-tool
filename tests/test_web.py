from io import BytesIO
from pathlib import Path
import time

from openpyxl import load_workbook

import excel_splitter.web.routes as routes_module
from excel_splitter.models import SplitSummary
from excel_splitter.web.app import create_app


def _app(tmp_path):
    return create_app(
        {
            "TESTING": True,
            "UPLOAD_DIR": tmp_path / "uploads",
            "DEFAULT_OUTPUT_DIR": tmp_path / "outputs",
        }
    )


def test_web_workflow_loads_configures_and_executes(sample_workbook, tmp_path):
    app = _app(tmp_path)
    client = app.test_client()

    with sample_workbook.open("rb") as source:
        loaded = client.post(
            "/api/load",
            data={"file": (BytesIO(source.read()), "人员数据.xlsx")},
            content_type="multipart/form-data",
        )
    assert loaded.status_code == 200
    payload = loaded.get_json()
    assert payload["sheets"] == ["人员", "项目", "说明"]
    assert payload["default_output_dir"] == str(tmp_path / "outputs")
    job_id = payload["job_id"]

    preview = client.post(
        "/api/preview",
        json={"job_id": job_id, "sheet_name": "人员", "start_row": 1, "max_rows": 2},
    )
    assert preview.status_code == 200
    assert preview.get_json()["suggested_header_row"] == 2
    assert preview.get_json()["start_row"] == 1
    assert preview.get_json()["end_row"] == 2
    assert preview.get_json()["total_rows"] == 5
    assert preview.get_json()["has_more"] is True

    next_page = client.post(
        "/api/preview",
        json={"job_id": job_id, "sheet_name": "人员", "start_row": 3, "max_rows": 2},
    )
    assert next_page.status_code == 200
    assert next_page.get_json()["start_row"] == 3
    assert next_page.get_json()["end_row"] == 4

    configs = [
        {
            "sheet_name": "人员",
            "header_row": 2,
            "split_column_idx": 1,
            "split_column_label": "A - 部门",
        },
        {
            "sheet_name": "项目",
            "header_row": 1,
            "split_column_idx": 1,
            "split_column_label": "A - 所属部门",
        },
    ]
    values = client.post(
        "/api/split-values", json={"job_id": job_id, "sheet_configs": configs}
    )
    assert values.status_code == 200
    assert values.get_json()["values"] == ["临床部", "研发部", "市场部"]

    executed = client.post(
        "/api/execute",
        json={
            "job_id": job_id,
            "sheet_configs": configs,
            "split_mode": "selected",
            "selected_split_values": ["临床部"],
            "output_dir": str(tmp_path / "custom-output"),
            "filename_template": "{original_name}_{split_value}",
            "overwrite": False,
        },
    )
    assert executed.status_code == 200
    result = executed.get_json()
    assert result["total_files"] == 2
    assert result["results"][0]["split_value"] == "临床部"
    assert [item["output_type"] for item in result["results"][0]["output_files"]] == [
        "formula",
        "values",
    ]

    formula_file = tmp_path / "custom-output" / "人员数据_临床部_公式版.xlsx"
    values_file = tmp_path / "custom-output" / "人员数据_临床部_结果值版.xlsx"
    assert values_file.exists()
    output = load_workbook(formula_file)
    assert output.sheetnames == ["人员", "项目"]
    output.close()

    for output_file in result["results"][0]["output_files"]:
        assert "download_url" not in output_file
        assert Path(output_file["output_file"]).is_file()
        assert output_file["output_type"] in ("formula", "values")


def test_web_returns_json_error_for_invalid_extension(tmp_path):
    client = _app(tmp_path).test_client()

    response = client.post(
        "/api/load",
        data={"file": (BytesIO(b"not excel"), "data.csv")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert ".xlsx" in response.get_json()["error"]


def test_web_executes_reference_and_linked_sheet_configuration(
    linked_workbook, tmp_path
):
    client = _app(tmp_path).test_client()
    with linked_workbook.open("rb") as source:
        loaded = client.post(
            "/api/load",
            data={"file": (BytesIO(source.read()), "关联拆分.xlsx")},
            content_type="multipart/form-data",
        )
    job_id = loaded.get_json()["job_id"]
    configs = [
        {
            "sheet_name": "人员归属",
            "header_row": 1,
            "mode": "reference",
            "split_column_idx": 1,
            "split_column_label": "A - 二级团队",
            "key_column_idx": 2,
            "key_column_label": "B - 姓名",
        },
        {
            "sheet_name": "工时明细",
            "header_row": 1,
            "mode": "linked",
            "split_column_idx": None,
            "key_column_idx": 1,
            "key_column_label": "A - 员工姓名",
        },
        {
            "sheet_name": "说明",
            "header_row": 1,
            "mode": "full",
            "split_column_idx": None,
            "key_column_idx": None,
        },
    ]

    values = client.post(
        "/api/split-values", json={"job_id": job_id, "sheet_configs": configs}
    )

    assert values.status_code == 200
    assert values.get_json()["values"] == ["团队甲", "团队乙"]

    executed = client.post(
        "/api/execute",
        json={
            "job_id": job_id,
            "sheet_configs": configs,
            "split_mode": "selected",
            "selected_split_values": ["团队甲"],
            "output_dir": str(tmp_path / "linked-output"),
            "overwrite": False,
        },
    )

    assert executed.status_code == 200
    result = executed.get_json()
    assert result["total_unmatched"] == 1
    assert result["results"][0]["sheet_rows"]["工时明细"] == 2
    assert result["results"][0]["unmatched_key_rows"]["工时明细"] == 1


def test_web_returns_complete_copy_value_when_all_sheets_are_full(
    linked_workbook, tmp_path
):
    client = _app(tmp_path).test_client()
    with linked_workbook.open("rb") as source:
        loaded = client.post(
            "/api/load",
            data={"file": (BytesIO(source.read()), "关联拆分.xlsx")},
            content_type="multipart/form-data",
        )
    configs = [
        {
            "sheet_name": "人员归属",
            "header_row": 1,
            "mode": "full",
            "split_column_idx": None,
            "key_column_idx": None,
        }
    ]

    values = client.post(
        "/api/split-values",
        json={"job_id": loaded.get_json()["job_id"], "sheet_configs": configs},
    )

    assert values.status_code == 200
    assert values.get_json()["values"] == ["完整表"]


def test_frontend_exposes_all_sheet_processing_modes(tmp_path):
    client = _app(tmp_path).test_client()

    page = client.get("/").get_data(as_text=True)
    script = client.get("/static/app.js").get_data(as_text=True)
    styles = client.get("/static/app.css").get_data(as_text=True)

    assert 'value="full" selected' in script
    assert 'id="select-all-sheets"' in page
    assert "全选" in page
    assert 'value="reference"' in script
    assert 'value="linked"' in script
    assert 'value="direct"' in script
    assert "不拆分，完整保留" in script
    assert 'id="unmatched-count"' in page
    assert ".config-controls label[hidden]" in styles
    assert "splitColumn.disabled = !usesSplitColumn" in script
    assert "Math.min(15, preview.total_rows)" in script
    assert 'class="btn-action"' in script
    assert '"open-file"' in script
    assert '"open-folder"' in script
    assert "打开文件" in script
    assert "打开所在文件夹" in script
    assert "加密输出文件" in page
    assert 'id="output-encrypt"' in page
    assert 'id="output-password"' in page
    assert 'id="password-modal"' in page
    assert 'id="password-input"' in page
    assert "文件已加密" in page
    assert "output_password" in script
    assert 'id="browse-output-dir"' in page
    assert 'id="execution-progress"' in page
    assert 'name="output-type" value="formula"' in page
    assert 'name="output-type" value="values"' in page
    assert 'id="selected-file-size"' in page
    assert "output_types" in script
    assert 'id="value-toolbar"' in page
    assert ".sheet-config { min-width: 0;" in styles
    assert ".preview-wrap { width: 100%; max-width: 100%; overflow: auto;" in styles


def test_web_can_choose_output_directory(tmp_path, monkeypatch):
    chosen = tmp_path / "chosen-output"
    monkeypatch.setattr(
        "excel_splitter.web.routes.choose_directory",
        lambda initial_dir: chosen,
    )
    client = _app(tmp_path).test_client()

    response = client.post(
        "/api/select-output-dir", json={"current_path": str(tmp_path)}
    )

    assert response.status_code == 200
    assert response.get_json() == {"selected": True, "path": str(chosen)}


def test_web_respects_selected_output_types(sample_workbook, tmp_path):
    client = _app(tmp_path).test_client()
    with sample_workbook.open("rb") as source:
        loaded = client.post(
            "/api/load",
            data={"file": (BytesIO(source.read()), "人员数据.xlsx")},
            content_type="multipart/form-data",
        )
    configs = [
        {
            "sheet_name": "人员",
            "header_row": 2,
            "split_column_idx": 1,
            "mode": "direct",
        }
    ]

    response = client.post(
        "/api/execute",
        json={
            "job_id": loaded.get_json()["job_id"],
            "sheet_configs": configs,
            "split_mode": "selected",
            "selected_split_values": ["临床部"],
            "output_types": ["values"],
            "output_dir": str(tmp_path / "values-only"),
        },
    )

    assert response.status_code == 200
    result = response.get_json()
    assert result["total_files"] == 1
    assert result["results"][0]["output_files"][0]["output_type"] == "values"


def test_web_rejects_empty_output_type_selection(sample_workbook, tmp_path):
    client = _app(tmp_path).test_client()
    with sample_workbook.open("rb") as source:
        loaded = client.post(
            "/api/load",
            data={"file": (BytesIO(source.read()), "人员数据.xlsx")},
            content_type="multipart/form-data",
        )

    response = client.post(
        "/api/execute",
        json={
            "job_id": loaded.get_json()["job_id"],
            "sheet_configs": [
                {
                    "sheet_name": "人员",
                    "header_row": 2,
                    "split_column_idx": 1,
                    "mode": "direct",
                }
            ],
            "output_types": [],
            "output_dir": str(tmp_path / "no-output"),
        },
    )

    assert response.status_code == 400
    assert "至少选择一种输出版本" in response.get_json()["error"]


def test_web_background_execute_reports_progress(sample_workbook, tmp_path):
    app = _app(tmp_path)
    client = app.test_client()
    with sample_workbook.open("rb") as source:
        loaded = client.post(
            "/api/load",
            data={"file": (BytesIO(source.read()), "人员数据.xlsx")},
            content_type="multipart/form-data",
        )
    configs = [
        {
            "sheet_name": "人员",
            "header_row": 2,
            "split_column_idx": 1,
            "mode": "direct",
        }
    ]

    started = client.post(
        "/api/execute",
        json={
            "job_id": loaded.get_json()["job_id"],
            "sheet_configs": configs,
            "split_mode": "selected",
            "selected_split_values": ["临床部"],
            "output_types": ["values"],
            "output_dir": str(tmp_path / "background-output"),
            "background": True,
        },
    )

    assert started.status_code == 202
    progress_url = started.get_json()["progress_url"]
    snapshots = []
    for _ in range(100):
        snapshot = client.get(progress_url).get_json()
        snapshots.append(snapshot)
        if snapshot["status"] in ("complete", "failed"):
            break
        time.sleep(0.01)

    assert snapshots[-1]["status"] == "complete"
    assert snapshots[-1]["progress"] == 100
    assert snapshots[-1]["message"]
    assert snapshots[-1]["result"]["total_files"] == 1

def test_web_reuses_cached_plan_only_for_matching_configs(
    sample_workbook, tmp_path, monkeypatch
):
    app = _app(tmp_path)
    client = app.test_client()
    with sample_workbook.open("rb") as source:
        loaded = client.post(
            "/api/load",
            data={"file": (BytesIO(source.read()), "data.xlsx")},
            content_type="multipart/form-data",
        )
    job_id = loaded.get_json()["job_id"]
    configs = [
        {
            "sheet_name": "人员",
            "header_row": 2,
            "split_column_idx": 1,
            "split_column_label": "A - 部门",
            "mode": "direct",
        }
    ]
    values = client.post(
        "/api/split-values",
        json={"job_id": job_id, "sheet_configs": configs},
    )
    assert values.status_code == 200
    record = app.extensions["excel_splitter_jobs"][job_id]
    cached_plan = record["_split_plan"]

    plans = []

    def capture_execute(self, job, progress_callback=None, plan=None):
        plans.append(plan)
        return SplitSummary(
            results=[],
            total_files=0,
            total_discarded=0,
        )

    monkeypatch.setattr(routes_module.SplitEngine, "execute", capture_execute)
    payload = {
        "job_id": job_id,
        "sheet_configs": configs,
        "split_mode": "selected",
        "selected_split_values": ["临床部"],
        "output_types": ["values"],
        "output_dir": str(tmp_path / "out"),
    }

    same = client.post("/api/execute", json=payload)
    assert same.status_code == 200
    assert plans[-1] is cached_plan

    changed_configs = [{**configs[0], "split_column_label": "changed"}]
    changed = client.post(
        "/api/execute",
        json={**payload, "sheet_configs": changed_configs},
    )
    assert changed.status_code == 200
    assert plans[-1] is None
