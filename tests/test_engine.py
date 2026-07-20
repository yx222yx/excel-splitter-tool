from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

import excel_splitter.engine as engine_module
from excel_splitter.engine import SplitEngine, _filter_linked_sheet, _filter_sheet
from excel_splitter.models import SheetConfig, SplitJob, SplitResult


class _Cell:
    def __init__(self, value):
        self.value = value


class _RecordingSheet:
    def __init__(self, max_row: int, values_by_row: dict[int, str | None]):
        self.max_row = max_row
        self.values_by_row = values_by_row
        self.deleted_rows: list[tuple[int, int]] = []

    def cell(self, row: int, column: int):
        return _Cell(self.values_by_row.get(row))

    def delete_rows(self, idx: int, amount: int = 1) -> None:
        self.deleted_rows.append((idx, amount))


def _output_file(result, output_type: str):
    return next(
        item.output_file for item in result.output_files if item.output_type == output_type
    )


def _job(sample_workbook: Path, output_dir: Path, mode: str = "all") -> SplitJob:
    return SplitJob(
        input_file=sample_workbook,
        output_dir=output_dir,
        sheet_configs=(
            SheetConfig("人员", 2, 1, "A - 部门"),
            SheetConfig("项目", 1, 1, "A - 所属部门"),
        ),
        split_mode=mode,
        selected_split_values=("临床部",) if mode == "selected" else (),
    )


def test_engine_exports_formula_and_values_workbooks_per_union_value(sample_workbook, tmp_path):
    summary = SplitEngine().execute(_job(sample_workbook, tmp_path / "输出"))

    assert summary.total_files == 6
    assert [result.split_value for result in summary.results] == [
        "临床部",
        "研发部",
        "市场部",
    ]
    assert summary.errors == []

    assert [item.output_type for item in summary.results[0].output_files] == [
        "formula",
        "values",
    ]

    clinical = load_workbook(
        _output_file(summary.results[0], "formula"), data_only=False
    )
    assert clinical.sheetnames == ["人员", "项目"]
    assert clinical["人员"].max_row == 3
    assert clinical["人员"]["A3"].value == "临床部"
    assert clinical["人员"]["D3"].value == "=C3*2"
    assert clinical["项目"].max_row == 2
    assert clinical["项目"]["A2"].value == "临床部"
    assert clinical["人员"].freeze_panes == "B3"
    assert clinical["人员"].column_dimensions["A"].width == 18
    assert "A1:D1" in {str(item) for item in clinical["人员"].merged_cells.ranges}
    clinical.close()

    clinical_values = load_workbook(
        _output_file(summary.results[0], "values"), data_only=False
    )
    assert clinical_values["人员"]["A3"].value == "临床部"
    assert clinical_values["人员"]["D3"].value is None
    clinical_values.close()


def test_union_value_missing_from_sheet_keeps_titles_and_header(sample_workbook, tmp_path):
    summary = SplitEngine().execute(_job(sample_workbook, tmp_path / "输出"))
    market_result = next(item for item in summary.results if item.split_value == "市场部")

    market = load_workbook(_output_file(market_result, "formula"))
    assert market["人员"].max_row == 2
    assert market["人员"]["A1"].value == "人员奖金明细"
    assert market["人员"]["A2"].value == "部门"
    assert market["项目"]["A2"].value == "市场部"
    market.close()


def test_selected_mode_exports_only_selected_values_and_reports_empty_rows(
    sample_workbook, tmp_path
):
    summary = SplitEngine().execute(
        _job(sample_workbook, tmp_path / "输出", mode="selected")
    )

    assert summary.total_files == 2
    assert summary.results[0].split_value == "临床部"
    assert summary.results[0].discarded_empty_rows == {"人员": 1, "项目": 0}
    assert summary.total_discarded == 1


def test_engine_filters_by_cached_result_but_preserves_formula(formula_workbook, tmp_path):
    job = SplitJob(
        input_file=formula_workbook,
        output_dir=tmp_path / "输出",
        sheet_configs=(SheetConfig("公式数据", 1, 1, "A - 部门公式"),),
        split_mode="selected",
        selected_split_values=("临床部",),
    )

    summary = SplitEngine().execute(job)

    formula_output = load_workbook(
        _output_file(summary.results[0], "formula"), data_only=False
    )
    assert formula_output["公式数据"].max_row == 2
    assert formula_output["公式数据"]["A2"].value == '=IF(1=1,"临床部","")'
    assert formula_output["公式数据"]["B2"].value == "张三"
    formula_output.close()

    values_output = load_workbook(
        _output_file(summary.results[0], "values"), data_only=False
    )
    assert values_output["公式数据"].max_row == 2
    assert values_output["公式数据"]["A2"].value == "临床部"
    assert values_output["公式数据"]["C2"].value == 2
    values_output.close()


def test_engine_uses_reference_keys_to_filter_linked_sheets_and_keeps_full_sheets(
    linked_workbook, tmp_path
):
    job = SplitJob(
        input_file=linked_workbook,
        output_dir=tmp_path / "输出",
        sheet_configs=(
            SheetConfig(
                "人员归属", 1, 1, mode="reference", key_column_idx=2
            ),
            SheetConfig("工时明细", 1, None, mode="linked", key_column_idx=1),
            SheetConfig("团队汇总", 1, 1, mode="direct"),
            SheetConfig("说明", 1, None, mode="full"),
        ),
        split_mode="selected",
        selected_split_values=("团队甲",),
    )

    summary = SplitEngine().execute(job)

    result = summary.results[0]
    assert result.sheet_rows == {
        "人员归属": 2,
        "工时明细": 2,
        "团队汇总": 1,
        "说明": 2,
    }
    assert result.discarded_empty_rows == {
        "人员归属": 0,
        "工时明细": 1,
        "团队汇总": 0,
        "说明": 0,
    }
    assert result.unmatched_key_rows == {
        "人员归属": 0,
        "工时明细": 1,
        "团队汇总": 0,
        "说明": 0,
    }
    assert summary.total_unmatched == 1

    output = load_workbook(_output_file(result, "formula"), data_only=False)
    assert [output["工时明细"].cell(row=row, column=1).value for row in range(2, 4)] == [
        "张三",
        "李四",
    ]
    assert output["说明"].max_row == 3
    output.close()


def test_engine_exports_one_complete_copy_when_all_sheets_are_full(
    linked_workbook, tmp_path
):
    job = SplitJob(
        input_file=linked_workbook,
        output_dir=tmp_path / "输出",
        sheet_configs=(
            SheetConfig("人员归属", 1, None, mode="full"),
            SheetConfig("说明", 1, None, mode="full"),
        ),
    )

    summary = SplitEngine().execute(job)

    assert summary.total_files == 2
    assert summary.results[0].split_value == "完整表"
    output = load_workbook(_output_file(summary.results[0], "formula"))
    assert output["人员归属"].max_row == 4
    assert output["说明"].max_row == 3
    output.close()


def test_engine_can_export_formula_version_only(sample_workbook, tmp_path):
    job = _job(sample_workbook, tmp_path / "输出", mode="selected")
    job = SplitJob(
        input_file=job.input_file,
        output_dir=job.output_dir,
        sheet_configs=job.sheet_configs,
        split_mode=job.split_mode,
        selected_split_values=job.selected_split_values,
        output_types=("formula",),
    )

    summary = SplitEngine().execute(job)

    assert summary.total_files == 1
    assert [item.output_type for item in summary.results[0].output_files] == [
        "formula"
    ]
    workbook = load_workbook(_output_file(summary.results[0], "formula"))
    assert workbook["人员"]["D3"].value == "=C3*2"
    workbook.close()


def test_engine_can_export_values_version_only(sample_workbook, tmp_path):
    job = _job(sample_workbook, tmp_path / "输出", mode="selected")
    job = SplitJob(
        input_file=job.input_file,
        output_dir=job.output_dir,
        sheet_configs=job.sheet_configs,
        split_mode=job.split_mode,
        selected_split_values=job.selected_split_values,
        output_types=("values",),
    )

    summary = SplitEngine().execute(job)

    assert summary.total_files == 1
    assert [item.output_type for item in summary.results[0].output_files] == [
        "values"
    ]
    workbook = load_workbook(_output_file(summary.results[0], "values"))
    assert workbook["人员"]["D3"].value is None
    workbook.close()


def test_engine_reports_monotonic_progress_to_completion(sample_workbook, tmp_path):
    events = []

    SplitEngine().execute(
        _job(sample_workbook, tmp_path / "输出", mode="selected"),
        progress_callback=lambda percent, message: events.append((percent, message)),
    )

    percentages = [percent for percent, _ in events]
    assert percentages[0] == 0
    assert percentages[-1] == 100
    assert percentages == sorted(percentages)
    assert all(message for _, message in events)


def test_filter_sheet_reaches_real_rows_above_virtual_empty_tail():
    sheet = _RecordingSheet(
        max_row=10006,
        values_by_row={2: "A", 3: "B", 4: "A"},
    )
    config = SheetConfig("Data", 1, 1)

    kept, _discarded, kept_rows = _filter_sheet(None, sheet, config, "A")

    assert kept == 2
    assert set(kept_rows) == {2, 4}


def test_filter_sheet_batches_contiguous_direct_deletions():
    values = {row: "B" for row in range(2, 102)}
    values[102] = "A"
    sheet = _RecordingSheet(max_row=102, values_by_row=values)
    config = SheetConfig("Data", 1, 1)

    kept, _discarded, kept_rows = _filter_sheet(None, sheet, config, "A")

    assert kept == 1
    assert kept_rows == [102]
    assert sheet.deleted_rows == [(2, 100)]


def test_filter_linked_sheet_reaches_real_rows_above_virtual_empty_tail():
    sheet = _RecordingSheet(
        max_row=10006,
        values_by_row={2: "A", 3: "B", 4: "A"},
    )
    config = SheetConfig("Detail", 1, None, mode="linked", key_column_idx=1)

    kept, _discarded, _unmatched, kept_rows = _filter_linked_sheet(
        None, sheet, config, {"A"}, {"A", "B"}
    )

    assert kept == 2
    assert set(kept_rows) == {2, 4}


def test_filter_linked_sheet_batches_contiguous_deletions():
    values = {row: "B" for row in range(2, 102)}
    values[102] = "A"
    sheet = _RecordingSheet(max_row=102, values_by_row=values)
    config = SheetConfig("Detail", 1, None, mode="linked", key_column_idx=1)

    kept, _discarded, _unmatched, kept_rows = _filter_linked_sheet(
        None, sheet, config, {"A"}, {"A", "B"}
    )

    assert kept == 1
    assert kept_rows == [102]
    assert sheet.deleted_rows == [(2, 100)]


def test_engine_maps_large_plan_progress_without_percent_regression(tmp_path):
    path = tmp_path / "large.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Group"])
    for row_index in range(1, 2101):
        sheet.append(["A" if row_index % 2 else "B"])
    workbook.save(path)
    workbook.close()

    events = []
    job = SplitJob(
        input_file=path,
        output_dir=tmp_path / "out",
        sheet_configs=(SheetConfig("Data", 1, 1),),
        split_mode="selected",
        selected_split_values=("A",),
        output_types=("values",),
    )

    SplitEngine().execute(
        job,
        progress_callback=lambda percent, message: events.append(
            (percent, message)
        ),
    )

    percentages = [percent for percent, _message in events]
    assert percentages == sorted(percentages)


def test_engine_reports_worker_progress_before_first_value_finishes(tmp_path):
    path = tmp_path / "progress.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Group"])
    sheet.append(["A"])
    sheet.append(["B"])
    workbook.save(path)
    workbook.close()

    class ReportingEngine(SplitEngine):
        def _export_value(self, job, split_value, plan, progress, **_kwargs):
            progress(0.25, f"inner {split_value}")
            return SplitResult(
                split_value=split_value,
                output_files=[],
                sheet_rows={},
                discarded_empty_rows={},
                unmatched_key_rows={},
                warnings=[],
            )

    events = []
    job = SplitJob(
        input_file=path,
        output_dir=tmp_path / "out",
        sheet_configs=(SheetConfig("Data", 1, 1),),
        split_mode="selected",
        selected_split_values=("A", "B"),
        output_types=("values",),
    )

    ReportingEngine().execute(
        job,
        progress_callback=lambda percent, message: events.append(
            (percent, message)
        ),
    )

    assert any(5 < percent < 50 for percent, _message in events)


def test_compact_filtered_sheet_moves_row_metadata_and_trims_ranges():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Group", "Name", "Amount"])
    sheet.append(["A", "First", 1])
    sheet.append(["B", "Removed", 2])
    sheet.append(["A", "Second", 3])
    sheet.row_dimensions[2].height = 22
    sheet.row_dimensions[3].height = 33
    sheet.row_dimensions[4].height = 44
    for row_index in range(5, 101):
        sheet.row_dimensions[row_index]
    sheet.auto_filter.ref = "A1:C100"
    sheet.add_table(Table(displayName="DataTable", ref="A1:C100"))
    config = SheetConfig("Data", 1, 1)

    _kept, _discarded, kept_rows = _filter_sheet(
        None, sheet, config, "A"
    )
    engine_module._compact_filtered_sheet(sheet, config, kept_rows)

    assert sheet.max_row == 3
    assert max(sheet.row_dimensions) == 3
    assert sheet.row_dimensions[2].height == 22
    assert sheet.row_dimensions[3].height == 44
    assert sheet.auto_filter.ref == "A1:C3"
    assert sheet.tables["DataTable"].ref == "A1:C3"
    workbook.close()

def test_engine_compacts_filtered_sheet_metadata_before_save(tmp_path):
    input_file = tmp_path / "metadata.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Group", "Name", "Amount"])
    sheet.append(["A", "First", 1])
    sheet.append(["B", "Removed", 2])
    sheet.append(["A", "Second", 3])
    sheet.row_dimensions[2].height = 22
    sheet.row_dimensions[3].height = 33
    sheet.row_dimensions[4].height = 44
    for row_index in range(5, 101):
        sheet.row_dimensions[row_index]
    sheet.auto_filter.ref = "A1:C100"
    sheet.add_table(Table(displayName="DataTable", ref="A1:C100"))
    workbook.save(input_file)
    workbook.close()

    job = SplitJob(
        input_file=input_file,
        output_dir=tmp_path / "out",
        sheet_configs=(SheetConfig("Data", 1, 1),),
        split_mode="selected",
        selected_split_values=("A",),
        output_types=("values",),
    )
    summary = SplitEngine().execute(job)

    output = load_workbook(_output_file(summary.results[0], "values"))
    compact = output["Data"]
    assert compact.max_row == 3
    assert max(compact.row_dimensions) == 3
    assert compact.row_dimensions[2].height == 22
    assert compact.row_dimensions[3].height == 44
    assert compact.auto_filter.ref == "A1:C3"
    assert compact.tables["DataTable"].ref == "A1:C3"
    output.close()


def test_fix_formula_references_does_not_scan_max_column_for_each_row(monkeypatch):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "Formula"
    sheet["A2"] = "=B4*2"
    config = SheetConfig("Data", 1, 1)
    worksheet_type = type(sheet)

    def fail_max_column(_sheet):
        raise AssertionError("max_column should not be accessed")

    monkeypatch.setattr(
        worksheet_type,
        "max_column",
        property(fail_max_column),
    )

    engine_module._fix_formula_references(
        workbook,
        (config,),
        {"Data": [4]},
    )

    assert sheet["A2"].value == "=B2*2"
    workbook.close()


def test_engine_uses_supplied_plan_without_rebuilding(
    sample_workbook, tmp_path, monkeypatch
):
    job = _job(sample_workbook, tmp_path / "out", mode="selected")
    plan = engine_module.build_split_plan(
        job.input_file,
        job.sheet_configs,
    )

    def fail_build(*_args, **_kwargs):
        raise AssertionError("plan should be reused")

    monkeypatch.setattr(engine_module, "build_split_plan", fail_build)

    summary = SplitEngine().execute(job, plan=plan)

    assert summary.errors == []
    assert summary.results[0].split_value == job.selected_split_values[0]


def test_select_process_workers_is_conservative(monkeypatch):
    monkeypatch.delenv("EXCEL_SPLITTER_WORKERS", raising=False)
    monkeypatch.setattr(engine_module.os, "cpu_count", lambda: 8)

    assert engine_module._select_process_workers(1, None) == 1
    assert engine_module._select_process_workers(10, None) == 2
    assert engine_module._select_process_workers(10, 3) == 3
    assert engine_module._select_process_workers(2, 9) == 2


def test_engine_process_pool_preserves_split_value_order(tmp_path, monkeypatch):
    input_file = tmp_path / "parallel.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Group"])
    sheet.append(["A"])
    sheet.append(["B"])
    workbook.save(input_file)
    workbook.close()

    submitted_workers = []

    class FakeFuture:
        def __init__(self, result):
            self._result = result

        def result(self):
            return self._result

    class FakeExecutor:
        def __init__(self, max_workers):
            submitted_workers.append(max_workers)

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def submit(self, function, *args):
            return FakeFuture(function(*args))

    def fake_export(_job, split_value, _plan, _file_bytes):
        return SplitResult(
            split_value=split_value,
            output_files=[],
            sheet_rows={},
            discarded_empty_rows={},
            unmatched_key_rows={},
            warnings=[],
        )

    monkeypatch.setattr(engine_module, "ProcessPoolExecutor", FakeExecutor)
    monkeypatch.setattr(engine_module, "_export_value_worker", fake_export)
    monkeypatch.setattr(
        engine_module,
        "as_completed",
        lambda futures: reversed(futures),
    )
    job = SplitJob(
        input_file=input_file,
        output_dir=tmp_path / "out",
        sheet_configs=(SheetConfig("Data", 1, 1),),
        split_mode="all",
        output_types=("values",),
    )
    plan = engine_module.SplitPlan(values=["A", "B"])

    summary = SplitEngine(process_workers=2).execute(job, plan=plan)

    assert submitted_workers == [2]
    assert [result.split_value for result in summary.results] == ["A", "B"]
    assert summary.errors == []
