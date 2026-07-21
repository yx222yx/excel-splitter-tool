from io import BytesIO
from pathlib import Path
import time

from openpyxl import Workbook, load_workbook

from excel_splitter.encryption import encrypt_file, is_encrypted
from excel_splitter.web.app import create_app


def _app(tmp_path):
    return create_app(
        {
            "TESTING": True,
            "UPLOAD_DIR": tmp_path / "uploads",
            "DEFAULT_OUTPUT_DIR": tmp_path / "outputs",
        }
    )


def _xlsx_bytes(sheets: dict[str, list[list]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _encrypted_xlsx_bytes(tmp_path: Path, name: str, sheets, password: str) -> bytes:
    path = tmp_path / name
    path.write_bytes(_xlsx_bytes(sheets))
    encrypt_file(path, password)
    data = path.read_bytes()
    path.unlink()
    return data


def _upload(client, files: list[tuple[bytes, str]], job_id: str | None = None):
    data: dict = {"files": [(BytesIO(content), name) for content, name in files]}
    if job_id:
        data["job_id"] = job_id
    return client.post("/api/merge/load", data=data, content_type="multipart/form-data")


def _simple_sheets() -> dict[str, list[list]]:
    return {"工时": [["姓名", "工时"], ["张三", 8]]}


def test_load_multiple_files_and_detects_encryption(tmp_path):
    client = _app(tmp_path).test_client()
    encrypted = _encrypted_xlsx_bytes(tmp_path, "加密.xlsx", _simple_sheets(), "pw123")

    response = _upload(
        client,
        [(_xlsx_bytes(_simple_sheets()), "部门A.xlsx"), (encrypted, "部门B.xlsx")],
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["has_encrypted"] is True
    assert [item["filename"] for item in payload["files"]] == [
        "部门A.xlsx",
        "部门B.xlsx",
    ]
    assert [item["encrypted"] for item in payload["files"]] == [False, True]
    assert all(item["size"] > 0 for item in payload["files"])

    appended = _upload(
        client,
        [(_xlsx_bytes(_simple_sheets()), "部门C.xlsx")],
        job_id=payload["job_id"],
    )
    assert appended.status_code == 200
    assert [item["filename"] for item in appended.get_json()["files"]] == [
        "部门A.xlsx",
        "部门B.xlsx",
        "部门C.xlsx",
    ]


def test_load_rejects_non_xlsx(tmp_path):
    client = _app(tmp_path).test_client()

    response = _upload(client, [(b"not excel", "data.csv")])

    assert response.status_code == 400
    assert ".xlsx" in response.get_json()["error"]


def test_unlock_with_unified_password_and_per_file_retry(tmp_path):
    client = _app(tmp_path).test_client()
    file_a = _encrypted_xlsx_bytes(tmp_path, "A.xlsx", _simple_sheets(), "统一密码")
    file_b = _encrypted_xlsx_bytes(tmp_path, "B.xlsx", _simple_sheets(), "独立密码")
    loaded = _upload(client, [(file_a, "部门A.xlsx"), (file_b, "部门B.xlsx")])
    job_id = loaded.get_json()["job_id"]
    file_b_id = loaded.get_json()["files"][1]["file_id"]

    response = client.post(
        "/api/merge/unlock", json={"job_id": job_id, "password": "统一密码"}
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["has_encrypted"] is True
    results = {item["filename"]: item for item in payload["results"]}
    assert results["部门A.xlsx"]["success"] is True
    assert results["部门B.xlsx"]["success"] is False
    assert "密码错误" in results["部门B.xlsx"]["error"]

    retry = client.post(
        "/api/merge/unlock",
        json={"job_id": job_id, "file_passwords": {file_b_id: "独立密码"}},
    )

    assert retry.status_code == 200
    assert retry.get_json()["has_encrypted"] is False
    assert retry.get_json()["results"][0]["success"] is True


def test_remove_and_reorder_files(tmp_path):
    app = _app(tmp_path)
    client = app.test_client()
    loaded = _upload(
        client,
        [
            (_xlsx_bytes(_simple_sheets()), "部门A.xlsx"),
            (_xlsx_bytes(_simple_sheets()), "部门B.xlsx"),
            (_xlsx_bytes(_simple_sheets()), "部门C.xlsx"),
        ],
    )
    job_id = loaded.get_json()["job_id"]
    files = loaded.get_json()["files"]
    id_a, id_b, id_c = (item["file_id"] for item in files)

    removed = client.post(
        "/api/merge/files/remove", json={"job_id": job_id, "file_id": id_b}
    )
    assert removed.status_code == 200
    assert [item["filename"] for item in removed.get_json()["files"]] == [
        "部门A.xlsx",
        "部门C.xlsx",
    ]

    reordered = client.post(
        "/api/merge/files/reorder",
        json={"job_id": job_id, "file_ids": [id_c, id_a]},
    )
    assert reordered.status_code == 200
    assert [item["filename"] for item in reordered.get_json()["files"]] == [
        "部门C.xlsx",
        "部门A.xlsx",
    ]

    invalid = client.post(
        "/api/merge/files/reorder",
        json={"job_id": job_id, "file_ids": [id_a]},
    )
    assert invalid.status_code == 400


def test_sheets_returns_union_in_first_file_order(tmp_path):
    client = _app(tmp_path).test_client()
    loaded = _upload(
        client,
        [
            (_xlsx_bytes({"工时": [["姓名"]], "汇总": [["指标"]]}), "部门A.xlsx"),
            (_xlsx_bytes({"项目": [["项目"]], "工时": [["姓名"]]}), "部门B.xlsx"),
        ],
    )

    response = client.post(
        "/api/merge/sheets", json={"job_id": loaded.get_json()["job_id"]}
    )

    assert response.status_code == 200
    assert response.get_json()["sheets"] == ["工时", "汇总", "项目"]


def test_plan_preview_reports_missing_and_extra_fields(tmp_path):
    app = _app(tmp_path)
    client = app.test_client()
    loaded = _upload(
        client,
        [
            (_xlsx_bytes({"工时": [["姓名", "工时"], ["张三", 8]]}), "部门A.xlsx"),
            (_xlsx_bytes({"工时": [["姓名", "项目"], ["王五", "X"]]}), "部门B.xlsx"),
        ],
    )
    job_id = loaded.get_json()["job_id"]
    configs = [{"sheet_name": "工时", "header_row": 1}]

    response = client.post(
        "/api/merge/plan", json={"job_id": job_id, "sheet_configs": configs}
    )

    assert response.status_code == 200
    sheet = response.get_json()["sheets"][0]
    assert sheet["union_headers"] == ["姓名", "工时", "项目"]
    assert sheet["base_file"] == "部门A.xlsx"
    assert sheet["missing_fields"]["部门B.xlsx"] == ["工时"]
    assert sheet["extra_fields"] == {"部门B.xlsx": ["项目"]}
    # plan 已缓存进 job，供 execute 复用
    record = app.extensions["excel_splitter_jobs"][job_id]
    assert record["_merge_plan"] is not None


def test_execute_end_to_end_sync(tmp_path):
    client = _app(tmp_path).test_client()
    loaded = _upload(
        client,
        [
            (_xlsx_bytes({"工时": [["姓名", "工时"], ["张三", 8], ["李四", 7]]}), "部门A.xlsx"),
            (_xlsx_bytes({"工时": [["工时", "姓名"], [6, "王五"]]}), "部门B.xlsx"),
        ],
    )
    job_id = loaded.get_json()["job_id"]

    response = client.post(
        "/api/merge/execute",
        json={
            "job_id": job_id,
            "sheet_configs": [{"sheet_name": "工时", "header_row": 1}],
            "output_dir": str(tmp_path / "合并输出"),
            "include_source_column": True,
            "background": False,
        },
    )

    assert response.status_code == 200
    result = response.get_json()
    assert result["errors"] == []
    assert result["total_rows"] == 3
    output_file = Path(result["output_file"])
    assert output_file.name == "部门A_合并结果.xlsx"
    assert output_file.is_file()

    workbook = load_workbook(output_file)
    rows = [list(row) for row in workbook["工时"].iter_rows(values_only=True)]
    workbook.close()
    assert rows == [
        ["姓名", "工时", "来源文件"],
        ["张三", 8, "部门A"],
        ["李四", 7, "部门A"],
        ["王五", 6, "部门B"],
    ]


def test_execute_defaults_output_dir_to_first_file_source_path(tmp_path):
    client = _app(tmp_path).test_client()
    source_dir = tmp_path / "原始目录"
    source_dir.mkdir()
    data = {
        "files": [
            (BytesIO(_xlsx_bytes(_simple_sheets())), "部门A.xlsx"),
            (BytesIO(_xlsx_bytes(_simple_sheets())), "部门B.xlsx"),
        ],
        "source_paths": [
            str(source_dir / "部门A.xlsx"),
            str(source_dir / "部门B.xlsx"),
        ],
    }
    loaded = client.post("/api/merge/load", data=data, content_type="multipart/form-data")
    job_id = loaded.get_json()["job_id"]

    response = client.post(
        "/api/merge/execute",
        json={
            "job_id": job_id,
            "sheet_configs": [{"sheet_name": "工时", "header_row": 1}],
            "background": False,
        },
    )

    assert response.status_code == 200
    output_file = Path(response.get_json()["output_file"])
    assert output_file.parent == source_dir
    assert output_file.name == "部门A_合并结果.xlsx"


def test_execute_with_output_password_encrypts_result(tmp_path):
    client = _app(tmp_path).test_client()
    loaded = _upload(
        client,
        [
            (_xlsx_bytes(_simple_sheets()), "部门A.xlsx"),
            (_xlsx_bytes(_simple_sheets()), "部门B.xlsx"),
        ],
    )

    response = client.post(
        "/api/merge/execute",
        json={
            "job_id": loaded.get_json()["job_id"],
            "sheet_configs": [{"sheet_name": "工时", "header_row": 1}],
            "output_dir": str(tmp_path / "加密输出"),
            "output_password": "输出密码",
            "background": False,
        },
    )

    assert response.status_code == 200
    output_file = Path(response.get_json()["output_file"])
    assert is_encrypted(output_file)


def test_execute_rejects_locked_files(tmp_path):
    client = _app(tmp_path).test_client()
    encrypted = _encrypted_xlsx_bytes(tmp_path, "加密.xlsx", _simple_sheets(), "pw")
    loaded = _upload(
        client, [(_xlsx_bytes(_simple_sheets()), "部门A.xlsx"), (encrypted, "部门B.xlsx")]
    )

    response = client.post(
        "/api/merge/execute",
        json={
            "job_id": loaded.get_json()["job_id"],
            "sheet_configs": [{"sheet_name": "工时", "header_row": 1}],
            "output_dir": str(tmp_path / "输出"),
            "background": False,
        },
    )

    assert response.status_code == 400
    assert "尚未解密" in response.get_json()["error"]


def test_background_execute_reports_progress(tmp_path):
    client = _app(tmp_path).test_client()
    loaded = _upload(
        client,
        [
            (_xlsx_bytes({"工时": [["姓名", "工时"], ["张三", 8]]}), "部门A.xlsx"),
            (_xlsx_bytes({"工时": [["姓名", "工时"], ["王五", 6]]}), "部门B.xlsx"),
        ],
    )
    job_id = loaded.get_json()["job_id"]

    started = client.post(
        "/api/merge/execute",
        json={
            "job_id": job_id,
            "sheet_configs": [{"sheet_name": "工时", "header_row": 1}],
            "output_dir": str(tmp_path / "后台输出"),
            "background": True,
        },
    )

    assert started.status_code == 202
    progress_url = started.get_json()["progress_url"]
    assert progress_url.endswith(f"/api/progress/{job_id}")

    snapshots = []
    for _ in range(200):
        snapshot = client.get(progress_url).get_json()
        snapshots.append(snapshot)
        if snapshot["status"] in ("complete", "failed"):
            break
        time.sleep(0.01)

    final = snapshots[-1]
    assert final["status"] == "complete"
    assert final["progress"] == 100
    result = final["result"]
    assert result["total_rows"] == 2
    assert result["results"][0]["sheet_name"] == "工时"
    assert result["results"][0]["source_rows"] == {"部门A.xlsx": 1, "部门B.xlsx": 1}
    assert Path(result["output_file"]).is_file()


def test_execute_skip_duplicate_sheets_param(tmp_path):
    client = _app(tmp_path).test_client()
    identical = {"汇总": [["指标"], [100], [200]]}
    loaded = _upload(
        client,
        [
            (_xlsx_bytes(identical), "部门A.xlsx"),
            (_xlsx_bytes(identical), "部门B.xlsx"),
        ],
    )
    job_id = loaded.get_json()["job_id"]
    configs = [{"sheet_name": "汇总", "header_row": 1}]

    default_run = client.post(
        "/api/merge/execute",
        json={
            "job_id": job_id,
            "sheet_configs": configs,
            "output_dir": str(tmp_path / "默认去重"),
            "output_filename": "默认去重.xlsx",
            "background": False,
        },
    )
    assert default_run.status_code == 200
    result = default_run.get_json()["results"][0]
    assert result["merged_rows"] == 2
    assert result["skipped_duplicates"] == {"部门B.xlsx": "部门A.xlsx"}

    disabled_run = client.post(
        "/api/merge/execute",
        json={
            "job_id": job_id,
            "sheet_configs": configs,
            "output_dir": str(tmp_path / "关闭去重"),
            "output_filename": "关闭去重.xlsx",
            "skip_duplicate_sheets": False,
            "background": False,
        },
    )
    assert disabled_run.status_code == 200
    result = disabled_run.get_json()["results"][0]
    assert result["merged_rows"] == 4
    assert result["skipped_duplicates"] == {}


def test_merge_preview_returns_first_rows_from_first_file(tmp_path):
    client = _app(tmp_path).test_client()
    loaded = _upload(
        client,
        [
            (_xlsx_bytes({"工时": [["姓名", "工时"], ["张三", 8], ["李四", 7]]}), "部门A.xlsx"),
            (_xlsx_bytes({"工时": [["姓名", "工时"], ["王五", 6]]}), "部门B.xlsx"),
        ],
    )
    job_id = loaded.get_json()["job_id"]

    response = client.post(
        "/api/merge/preview", json={"job_id": job_id, "sheet_name": "工时"}
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["source_file"] == "部门A.xlsx"
    assert payload["total_rows"] == 3
    assert payload["rows"] == [["姓名", "工时"], ["张三", 8], ["李四", 7]]

    missing = client.post(
        "/api/merge/preview", json={"job_id": job_id, "sheet_name": "不存在"}
    )
    assert missing.status_code == 400
    assert "没有文件包含 sheet" in missing.get_json()["error"]


def test_merge_plan_and_execute_with_identical_sheet(tmp_path):
    client = _app(tmp_path).test_client()
    loaded = _upload(
        client,
        [
            (
                _xlsx_bytes({
                    "汇总": [["指标"], [100]],
                    "工时": [["姓名"], ["张三"]],
                }),
                "部门A.xlsx",
            ),
            (
                _xlsx_bytes({
                    "汇总": [["指标"], [100]],
                    "工时": [["姓名"], ["王五"]],
                }),
                "部门B.xlsx",
            ),
        ],
    )
    job_id = loaded.get_json()["job_id"]
    configs = [
        {"sheet_name": "汇总", "header_row": 1, "identical": True},
        {"sheet_name": "工时", "header_row": 1},
    ]

    plan = client.post("/api/merge/plan", json={"job_id": job_id, "sheet_configs": configs})
    assert plan.status_code == 200
    sheets = plan.get_json()["sheets"]
    assert sheets[0] == {"sheet_name": "汇总", "identical": True}
    assert sheets[1]["sheet_name"] == "工时"
    assert sheets[1]["union_headers"] == ["姓名"]

    executed = client.post(
        "/api/merge/execute",
        json={
            "job_id": job_id,
            "sheet_configs": configs,
            "output_dir": str(tmp_path / "输出"),
            "background": False,
        },
    )
    assert executed.status_code == 200
    result = executed.get_json()
    by_sheet = {item["sheet_name"]: item for item in result["results"]}
    assert by_sheet["汇总"]["source_rows"] == {"部门A.xlsx": 1}
    assert by_sheet["工时"]["source_rows"] == {"部门A.xlsx": 1, "部门B.xlsx": 1}

    workbook = load_workbook(Path(result["output_file"]))
    rows = [list(row) for row in workbook["汇总"].iter_rows(values_only=True)]
    workbook.close()
    assert rows == [["指标"], [100]]
