from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference

from excel_splitter.merge_engine import MergeEngine
from excel_splitter.merge_models import MergeJob, MergeSheetConfig
from excel_splitter.web.app import create_app


def _sub_block_workbook(path: Path, rows: list[list]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "概览"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _files(tmp_path: Path):
    # A（模板）：主表 + 小表1（缺小表2）
    file_a = _sub_block_workbook(
        tmp_path / "部门A.xlsx",
        [
            ["团队", "指标"],
            ["甲", 1],
            [None, None],
            ["负责人", "人数"],
            ["甲", 10],
        ],
    )
    workbook = load_workbook(file_a)
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

    chart = BarChart()
    sheet = workbook["概览"]
    chart.add_data(Reference(sheet, min_col=2, min_row=2, max_row=2), titles_from_data=False)
    # 锚点落在小表区（0 基第 4 行 = 1 基第 5 行，即小表1数据行处）
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=3, colOff=0, row=4, rowOff=0)
    anchor.to = AnchorMarker(col=8, colOff=0, row=10, rowOff=0)
    chart.anchor = anchor
    sheet.add_chart(chart)
    workbook.save(file_a)
    workbook.close()
    # B：主表 + 小表1（不同行）+ 小表2
    file_b = _sub_block_workbook(
        tmp_path / "部门B.xlsx",
        [
            ["团队", "指标"],
            ["乙", 2],
            [None, None],
            ["负责人", "人数"],
            ["乙", 20],
            [None, None],
            ["组长", "项目"],
            ["乙", "P1"],
        ],
    )
    # C：主表 + 小表1（带公式）+ 小表2（与 B 完全相同）
    file_c = _sub_block_workbook(
        tmp_path / "部门C.xlsx",
        [
            ["团队", "指标"],
            ["丙", 3],
            [None, None],
            ["负责人", "人数"],
            ["丙", "=B5*10"],
            [None, None],
            ["组长", "项目"],
            ["乙", "P1"],
        ],
    )
    return file_a, file_b, file_c


def _read(path: Path):
    workbook = load_workbook(path)
    rows = [list(row) for row in workbook["概览"].iter_rows(values_only=True)]
    return workbook, rows


def test_sub_blocks_reinserted_after_main_merge(tmp_path):
    file_a, file_b, file_c = _files(tmp_path)
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b, file_c),
        output_file=output,
        sheet_configs=(MergeSheetConfig("概览", has_sub_blocks=True),),
    )
    summary = MergeEngine().execute(job)

    assert summary.errors == []
    workbook, rows = _read(output)
    # 主表连续合并 → 空行 → 小表1（表头 + 三文件拼接）→ 空行 → 小表2（表头 + B 一份）
    assert [row[0] for row in rows] == [
        "团队", "甲", "乙", "丙",
        None, "负责人", "甲", "乙", "丙",
        None, "组长", "乙",
    ]
    # C 小表1 的公式按目标行平移
    assert workbook["概览"]["B9"].value == "=B9*10"
    # C 的小表2 与 B 完全相同 → 只保留一份并记警告
    assert any("完全相同" in w for w in summary.results[0].warnings)
    # 图表锚点下移到最终内容末行（第 12 行）之后
    anchor = workbook["概览"]._charts[0].anchor
    assert anchor._from.row + 1 == 13
    assert anchor.to.row + 1 == 19
    workbook.close()


def test_sub_block_source_rows_counted(tmp_path):
    file_a, file_b, file_c = _files(tmp_path)
    output = tmp_path / "合并结果.xlsx"
    summary = MergeEngine().execute(
        MergeJob(
            input_files=(file_a, file_b, file_c),
            output_file=output,
            sheet_configs=(MergeSheetConfig("概览", has_sub_blocks=True),),
        )
    )

    source_rows = summary.results[0].source_rows
    # 主表 1 行 + 小表1 各 1 行；B 另有小表2 一行
    assert source_rows == {"部门A.xlsx": 2, "部门B.xlsx": 3, "部门C.xlsx": 2}


def _app(tmp_path):
    return create_app(
        {
            "TESTING": True,
            "UPLOAD_DIR": tmp_path / "uploads",
            "DEFAULT_OUTPUT_DIR": tmp_path / "outputs",
        }
    )


def test_merge_sheet_blocks_endpoint(tmp_path):
    client = _app(tmp_path).test_client()
    file_a, file_b, _ = _files(tmp_path)
    data = {
        "files": [
            (BytesIO(file_a.read_bytes()), "部门A.xlsx"),
            (BytesIO(file_b.read_bytes()), "部门B.xlsx"),
        ]
    }
    loaded = client.post("/api/merge/load", data=data, content_type="multipart/form-data")
    job_id = loaded.get_json()["job_id"]

    response = client.post(
        "/api/merge/sheet-blocks",
        json={"job_id": job_id, "sheet_name": "概览", "header_row": 1},
    )
    assert response.status_code == 200
    files = response.get_json()["files"]
    assert len(files) == 2
    a_blocks = files[0]["blocks"]
    assert [(b["header_row"], b["data_start"], b["data_end"]) for b in a_blocks] == [
        (1, 2, 2),
        (4, 5, 5),
    ]
    assert a_blocks[1]["header_preview"] == "负责人 / 人数"
    assert len(files[1]["blocks"]) == 3

    missing = client.post(
        "/api/merge/sheet-blocks",
        json={"job_id": job_id, "sheet_name": "不存在", "header_row": 1},
    )
    assert missing.status_code == 400


def test_execute_passes_has_sub_blocks(tmp_path):
    client = _app(tmp_path).test_client()
    file_a, file_b, _ = _files(tmp_path)
    data = {
        "files": [
            (BytesIO(file_a.read_bytes()), "部门A.xlsx"),
            (BytesIO(file_b.read_bytes()), "部门B.xlsx"),
        ]
    }
    loaded = client.post("/api/merge/load", data=data, content_type="multipart/form-data")
    job_id = loaded.get_json()["job_id"]

    executed = client.post(
        "/api/merge/execute",
        json={
            "job_id": job_id,
            "sheet_configs": [
                {"sheet_name": "概览", "header_row": 1, "has_sub_blocks": True}
            ],
            "output_dir": str(tmp_path / "输出"),
            "background": False,
        },
    )
    assert executed.status_code == 200
    output_file = Path(executed.get_json()["output_file"])
    workbook = load_workbook(output_file)
    rows = [row[0] for row in workbook["概览"].iter_rows(values_only=True)]
    workbook.close()
    assert rows == [
        "团队", "甲", "乙",
        None, "负责人", "甲", "乙",
        None, "组长", "乙",
    ]
