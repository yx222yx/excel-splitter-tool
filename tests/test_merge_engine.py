from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from excel_splitter.merge_engine import MergeEngine
from excel_splitter.merge_models import MergeJob, MergeSheetConfig
from excel_splitter.merge_planning import build_merge_plan


def _make_workbook(path: Path, sheets: dict[str, list[list]]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _read_rows(path: Path, sheet_name: str) -> list[list]:
    workbook = load_workbook(path)
    rows = [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
    workbook.close()
    return rows


def test_basic_merge_two_files(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时"], ["张三", 8], ["李四", 7]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "工时"], ["王五", 6]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    summary = MergeEngine().execute(job)

    assert summary.errors == []
    assert summary.total_rows == 3
    result = summary.results[0]
    assert result.sheet_name == "工时"
    assert result.merged_rows == 3
    assert result.source_rows == {"部门A.xlsx": 2, "部门B.xlsx": 1}
    assert _read_rows(output, "工时") == [
        ["姓名", "工时"],
        ["张三", 8],
        ["李四", 7],
        ["王五", 6],
    ]


def test_merge_aligns_by_header_name_when_column_order_differs(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "部门", "工时"], ["张三", "临床部", 8]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["工时", "姓名", "部门"], [6, "王五", "研发部"]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    summary = MergeEngine().execute(job)

    assert summary.errors == []
    assert _read_rows(output, "工时") == [
        ["姓名", "部门", "工时"],
        ["张三", "临床部", 8],
        ["王五", "研发部", 6],
    ]


def test_merge_unions_fields_and_reports_missing_and_extra(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时"], ["张三", 8]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "项目"], ["王五", "项目X"]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    summary = MergeEngine().execute(job)

    result = summary.results[0]
    # 缺失字段相对并集计算：B 缺「工时」，A 也缺 B 多出的「项目」
    assert result.missing_fields == {
        "部门A.xlsx": ["项目"],
        "部门B.xlsx": ["工时"],
    }
    assert result.extra_fields == {"部门B.xlsx": ["项目"]}
    assert any("部门B.xlsx" in w and "工时" in w for w in summary.warnings)
    assert _read_rows(output, "工时") == [
        ["姓名", "工时", "项目"],
        ["张三", 8, None],
        ["王五", None, "项目X"],
    ]


def test_merge_skips_file_missing_sheet_with_warning(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {
            "工时": [["姓名"], ["张三"]],
            "汇总": [["指标"], [10]],
        },
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名"], ["王五"]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"), MergeSheetConfig("汇总")),
    )

    summary = MergeEngine().execute(job)

    assert summary.errors == []
    assert any("部门B.xlsx" in w and "汇总" in w for w in summary.warnings)
    assert _read_rows(output, "汇总") == [["指标"], [10]]
    assert _read_rows(output, "工时") == [["姓名"], ["张三"], ["王五"]]


def test_merge_keeps_title_area_from_first_file_when_header_row_late(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["2026 年 7 月工时汇总", None], ["姓名", "工时"], ["张三", 8]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["部门B 自己的标题", None], ["姓名", "工时"], ["王五", 6]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时", header_row=2),),
    )

    summary = MergeEngine().execute(job)

    assert summary.errors == []
    assert _read_rows(output, "工时") == [
        ["2026 年 7 月工时汇总", None],
        ["姓名", "工时"],
        ["张三", 8],
        ["王五", 6],
    ]


def test_merge_skips_blank_rows(tmp_path):
    # 模板（第一个文件）的行照抄不动，空行保留；追加文件的整行皆空才跳过
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时"], ["张三", 8], ["李四", 7]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "工时"], ["王五", 6], [None, None], [" ", None], ["赵六", 5]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    summary = MergeEngine().execute(job)

    assert summary.total_rows == 4
    assert _read_rows(output, "工时") == [
        ["姓名", "工时"],
        ["张三", 8],
        ["李四", 7],
        ["王五", 6],
        ["赵六", 5],
    ]


def test_merge_adds_source_column_when_enabled(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名"], ["张三"]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名"], ["王五"]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
        include_source_column=True,
    )

    summary = MergeEngine().execute(job)

    assert summary.errors == []
    assert _read_rows(output, "工时") == [
        ["姓名", "来源文件"],
        ["张三", "部门A"],
        ["王五", "部门B"],
    ]


def test_merge_reuses_provided_plan(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时"], ["张三", 8]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["工时", "姓名"], [6, "王五"]]},
    )
    configs = (MergeSheetConfig("工时"),)
    plan = build_merge_plan((file_a, file_b), configs)
    assert plan.sheets[0].union_headers == ["姓名", "工时"]
    assert plan.sheets[0].base_file == "部门A.xlsx"
    assert plan.sheets[0].missing_files == []

    output_with_plan = tmp_path / "带plan.xlsx"
    output_without_plan = tmp_path / "不带plan.xlsx"
    summary_with_plan = MergeEngine().execute(
        MergeJob(
            input_files=(file_a, file_b),
            output_file=output_with_plan,
            sheet_configs=configs,
        ),
        plan=plan,
    )
    summary_without_plan = MergeEngine().execute(
        MergeJob(
            input_files=(file_a, file_b),
            output_file=output_without_plan,
            sheet_configs=configs,
        )
    )

    assert summary_with_plan.total_rows == summary_without_plan.total_rows == 2
    assert _read_rows(output_with_plan, "工时") == _read_rows(
        output_without_plan, "工时"
    )


def test_merge_reports_progress_in_order(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名"], ["张三"]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名"], ["王五"]]},
    )
    events: list[tuple[int, str]] = []
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=tmp_path / "合并结果.xlsx",
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    MergeEngine().execute(job, progress_callback=lambda p, m: events.append((p, m)))

    percents = [percent for percent, _ in events]
    assert percents[0] == 0
    assert percents[-1] == 100
    assert percents == sorted(percents)


def test_merge_copies_header_style_and_column_width_from_first_file(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时"], ["张三", 8]]},
    )
    workbook = load_workbook(file_a)
    sheet = workbook["工时"]
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 21
    workbook.save(file_a)
    workbook.close()
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "工时"], ["王五", 6]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    MergeEngine().execute(job)

    workbook = load_workbook(output)
    merged = workbook["工时"]
    assert merged["A1"].font.bold
    assert merged.column_dimensions["A"].width == 21
    workbook.close()


def _merge_job(files, output, sheet_names=("汇总",), **overrides):
    kwargs = {
        "input_files": tuple(files),
        "output_file": output,
        "sheet_configs": tuple(MergeSheetConfig(name) for name in sheet_names),
    }
    kwargs.update(overrides)
    return MergeJob(**kwargs)


def test_merge_skips_identical_sheet_by_default(tmp_path):
    identical = {"汇总": [["指标", "说明"], [100, "相同内容"], [200, "完全一致"]]}
    file_a = _make_workbook(tmp_path / "部门A.xlsx", identical)
    file_b = _make_workbook(tmp_path / "部门B.xlsx", identical)
    output = tmp_path / "合并结果.xlsx"

    summary = MergeEngine().execute(_merge_job([file_a, file_b], output))

    result = summary.results[0]
    assert result.merged_rows == 2
    assert result.source_rows == {"部门A.xlsx": 2}
    assert result.skipped_duplicates == {"部门B.xlsx": "部门A.xlsx"}
    assert any("部门B.xlsx" in w and "完全相同" in w for w in summary.warnings)
    assert _read_rows(output, "汇总") == [
        ["指标", "说明"],
        [100, "相同内容"],
        [200, "完全一致"],
    ]


def test_merge_writes_duplicates_when_skip_disabled(tmp_path):
    identical = {"汇总": [["指标"], [100], [200]]}
    file_a = _make_workbook(tmp_path / "部门A.xlsx", identical)
    file_b = _make_workbook(tmp_path / "部门B.xlsx", identical)
    output = tmp_path / "合并结果.xlsx"

    summary = MergeEngine().execute(
        _merge_job([file_a, file_b], output, skip_duplicate_sheets=False)
    )

    result = summary.results[0]
    assert result.merged_rows == 4
    assert result.skipped_duplicates == {}
    assert _read_rows(output, "汇总") == [["指标"], [100], [200], [100], [200]]


def test_merge_skips_only_the_identical_sheet(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {
            "汇总": [["指标"], [100]],
            "工时": [["姓名"], ["张三"]],
        },
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {
            "汇总": [["指标"], [100]],
            "工时": [["姓名"], ["王五"]],
        },
    )
    output = tmp_path / "合并结果.xlsx"

    summary = MergeEngine().execute(
        _merge_job([file_a, file_b], output, sheet_names=("汇总", "工时"))
    )

    results = {r.sheet_name: r for r in summary.results}
    assert results["汇总"].skipped_duplicates == {"部门B.xlsx": "部门A.xlsx"}
    assert results["汇总"].merged_rows == 1
    assert results["工时"].skipped_duplicates == {}
    assert results["工时"].merged_rows == 2


def test_merge_keeps_first_of_three_files_when_two_identical(tmp_path):
    file_a = _make_workbook(tmp_path / "部门A.xlsx", {"汇总": [["指标"], [100]]})
    file_b = _make_workbook(tmp_path / "部门B.xlsx", {"汇总": [["指标"], [200]]})
    file_c = _make_workbook(tmp_path / "部门C.xlsx", {"汇总": [["指标"], [200]]})
    output = tmp_path / "合并结果.xlsx"

    summary = MergeEngine().execute(_merge_job([file_a, file_b, file_c], output))

    result = summary.results[0]
    assert result.merged_rows == 2
    assert result.source_rows == {"部门A.xlsx": 1, "部门B.xlsx": 1}
    assert result.skipped_duplicates == {"部门C.xlsx": "部门B.xlsx"}
    assert _read_rows(output, "汇总") == [["指标"], [100], [200]]


def test_merge_does_not_treat_different_column_order_as_duplicate(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx", {"工时": [["姓名", "工时"], ["张三", 8]]}
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx", {"工时": [["工时", "姓名"], [8, "张三"]]}
    )
    output = tmp_path / "合并结果.xlsx"

    summary = MergeEngine().execute(_merge_job([file_a, file_b], output, sheet_names=("工时",)))

    result = summary.results[0]
    assert result.skipped_duplicates == {}
    assert result.merged_rows == 2


def test_merge_preserves_data_cell_styles(tmp_path):
    from datetime import datetime

    from openpyxl.styles import Font, PatternFill

    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时", "日期"], ["张三", 8, datetime(2026, 7, 1)]]},
    )
    workbook = load_workbook(file_a)
    sheet = workbook["工时"]
    sheet["A2"].font = Font(bold=True, color="FF0000")
    sheet["A2"].fill = PatternFill("solid", fgColor="FFFF00")
    sheet["B2"].number_format = "#,##0.0"
    sheet["C2"].number_format = "yyyy-mm-dd"
    workbook.save(file_a)
    workbook.close()

    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "工时", "日期"], ["李四", 7, None]]},
    )
    workbook = load_workbook(file_b)
    workbook["工时"]["A2"].font = Font(italic=True)
    workbook.save(file_b)
    workbook.close()

    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )
    MergeEngine().execute(job)

    workbook = load_workbook(output)
    sheet = workbook["工时"]
    assert sheet["A2"].font.bold
    assert sheet["A2"].font.color.rgb.endswith("FF0000")
    assert sheet["A2"].fill.fgColor.rgb.endswith("FFFF00")
    assert sheet["B2"].number_format == "#,##0.0"
    assert sheet["C2"].number_format == "yyyy-mm-dd"
    # 第二个文件的数据行保留自己的样式
    assert sheet["A3"].font.italic
    assert not sheet["A3"].font.bold
    workbook.close()


def test_merge_preserves_title_merge_freeze_and_row_height(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {
            "工时": [
                ["2026 年 7 月工时汇总", None, None],
                ["姓名", "项目", "工时"],
                ["张三", "项目甲", 8],
            ]
        },
    )
    workbook = load_workbook(file_a)
    sheet = workbook["工时"]
    sheet.merge_cells("A1:C1")
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 22
    sheet.freeze_panes = "A3"
    workbook.save(file_a)
    workbook.close()

    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {
            "工时": [
                ["部门B 的标题", None, None],
                ["姓名", "项目", "工时"],
                ["李四", "项目乙", 7],
            ]
        },
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时", header_row=2),),
    )
    summary = MergeEngine().execute(job)

    assert summary.errors == []
    workbook = load_workbook(output)
    sheet = workbook["工时"]
    assert sheet["A1"].value == "2026 年 7 月工时汇总"
    assert sheet["A1"].font.bold
    assert "A1:C1" in {str(item) for item in sheet.merged_cells.ranges}
    assert sheet.row_dimensions[1].height == 30
    assert sheet.row_dimensions[2].height == 22
    assert sheet.freeze_panes == "A3"
    assert sheet["A4"].value == "李四"
    workbook.close()


def test_merge_preserves_title_rows_when_header_row_is_three(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {
            "工时": [
                ["2026 年 7 月工时汇总", None, None],
                ["编制：人事部", None, "统计区间：7.1-7.31"],
                ["姓名", "项目", "工时"],
                ["张三", "项目甲", 8],
            ]
        },
    )
    workbook = load_workbook(file_a)
    sheet = workbook["工时"]
    sheet.merge_cells("A1:C1")
    sheet.merge_cells("A2:B2")
    sheet.row_dimensions[1].height = 30
    workbook.save(file_a)
    workbook.close()

    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {
            "工时": [
                ["部门B 自己的标题", None, None],
                ["部门B 的副标题", None, None],
                ["姓名", "项目", "工时"],
                ["李四", "项目乙", 7],
            ]
        },
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时", header_row=3),),
    )
    summary = MergeEngine().execute(job)

    assert summary.errors == []
    rows = _read_rows(output, "工时")
    assert rows == [
        ["2026 年 7 月工时汇总", None, None],
        ["编制：人事部", None, "统计区间：7.1-7.31"],
        ["姓名", "项目", "工时"],
        ["张三", "项目甲", 8],
        ["李四", "项目乙", 7],
    ]
    workbook = load_workbook(output)
    sheet = workbook["工时"]
    assert {str(item) for item in sheet.merged_cells.ranges} == {"A1:C1", "A2:B2"}
    assert sheet.row_dimensions[1].height == 30
    workbook.close()


def test_merge_keeps_template_formulas_untouched(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时", "双倍"], ["张三", 8, "=B2*2"], ["李四", 7, "=B3*2"]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "工时", "双倍"], ["王五", 6, "=B2*2"]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    MergeEngine().execute(job)

    workbook = load_workbook(output)
    sheet = workbook["工时"]
    # 模板自己的公式原文不动
    assert sheet["C2"].value == "=B2*2"
    assert sheet["C3"].value == "=B3*2"
    workbook.close()


def test_merge_translates_appended_formulas(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时", "公式"], ["张三", 8, None], ["李四", 7, None]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {
            "工时": [
                ["姓名", "工时", "公式"],
                ["王五", 6, "=B2*2"],
                ["赵六", 5, "=$B$2+B3"],
                ["孙七", 4, "=汇总!A2*2"],
            ],
            "汇总": [["指标"], [100]],
        },
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"), MergeSheetConfig("汇总")),
    )

    MergeEngine().execute(job)

    workbook = load_workbook(output)
    sheet = workbook["工时"]
    # B 的数据行追加到第 4-6 行：相对引用平移、绝对引用不变、跨 sheet 引用也平移
    assert sheet["C4"].value == "=B4*2"
    assert sheet["C5"].value == "=$B$2+B5"
    assert sheet["C6"].value == "=汇总!A4*2"
    workbook.close()


def test_identical_sheet_keeps_template_and_never_reads_other_files(tmp_path, monkeypatch):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"汇总": [["指标"], [100], [200]], "工时": [["姓名"], ["张三"]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"汇总": [["指标"], [100], [200]], "工时": [["姓名"], ["王五"]]},
    )
    opened: list[Path] = []

    import excel_splitter.merge_engine as engine_module

    original_loader = engine_module.load_workbook_with_warnings

    def recording_loader(path, **kwargs):
        opened.append(Path(getattr(path, "name", path)))
        return original_loader(path, **kwargs)

    monkeypatch.setattr(engine_module, "load_workbook_with_warnings", recording_loader)

    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(
            MergeSheetConfig("汇总", identical=True),
            MergeSheetConfig("工时"),
        ),
        skip_duplicate_sheets=False,  # 关闭指纹去重后，B 只应为「工时」被读一次
    )
    summary = MergeEngine().execute(job)

    results = {r.sheet_name: r for r in summary.results}
    assert results["汇总"].source_rows == {"部门A.xlsx": 2}
    assert results["工时"].source_rows == {"部门A.xlsx": 1, "部门B.xlsx": 1}
    # identical 的「汇总」对其他文件零读取：B 仅因追加「工时」被读取一次
    reads_of_b = [
        p for p in opened
        if isinstance(p, Path) and p.name == "部门B.xlsx"
    ]
    assert len(reads_of_b) == 1
    assert _read_rows(output, "汇总") == [["指标"], [100], [200]]


def test_identical_sheet_copied_from_later_file_when_template_lacks_it(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名"], ["张三"]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"汇总": [["指标", "说明"], [100, "完整拷贝"]], "工时": [["姓名"], ["王五"]]},
    )
    workbook = load_workbook(file_b)
    workbook["汇总"].column_dimensions["A"].width = 23
    workbook.save(file_b)
    workbook.close()

    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(
            MergeSheetConfig("汇总", identical=True),
            MergeSheetConfig("工时"),
        ),
    )
    summary = MergeEngine().execute(job)

    results = {r.sheet_name: r for r in summary.results}
    assert results["汇总"].source_rows == {"部门B.xlsx": 1}
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["汇总", "工时"]
    sheet = workbook["汇总"]
    assert sheet["A1"].value == "指标"
    assert sheet["B2"].value == "完整拷贝"
    assert sheet.column_dimensions["A"].width == 23
    workbook.close()


def test_header_row_appears_only_once_with_three_files(tmp_path):
    files = [
        _make_workbook(
            tmp_path / f"部门{letter}.xlsx",
            {"工时": [["姓名", "工时"], [f"员工{letter}", index + 5]]},
        )
        for index, letter in enumerate("ABC")
    ]
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=tuple(files),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    MergeEngine().execute(job)

    workbook = load_workbook(output)
    column_a = [
        row[0] for row in workbook["工时"].iter_rows(values_only=True)
    ]
    workbook.close()
    assert column_a.count("姓名") == 1
    assert column_a == ["姓名", "员工A", "员工B", "员工C"]


def test_appended_formulas_translate_independently_per_file(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "单价", "数量", "金额"], ["张三", 10, 2, "=B2*C2"]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "单价", "数量", "金额"], ["王五", 20, 3, "=B2*C2"]]},
    )
    file_c = _make_workbook(
        tmp_path / "部门C.xlsx",
        {"工时": [["姓名", "单价", "数量", "金额"], ["赵六", 30, 4, "=B2*C2"]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b, file_c),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    MergeEngine().execute(job)

    workbook = load_workbook(output)
    sheet = workbook["工时"]
    # 每个文件的公式按自己的源坐标平移，互不串行
    assert sheet["D2"].value == "=B2*C2"  # 模板原文不动
    assert sheet["D3"].value == "=B3*C3"
    assert sheet["D4"].value == "=B4*C4"
    workbook.close()


def test_appended_range_formula_translates(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "一", "二", "三", "合计"], ["张三", 1, 2, 3, "=SUM(B2:D2)"]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "一", "二", "三", "合计"], ["王五", 4, 5, 6, "=SUM(B2:D2)"]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    MergeEngine().execute(job)

    workbook = load_workbook(output)
    assert workbook["工时"]["E3"].value == "=SUM(B3:D3)"
    workbook.close()


def test_formula_referencing_above_header_translates_and_warns(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时", "公式"], ["张三", 8, None], ["李四", 7, None], ["王五", 6, None]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"工时": [["姓名", "工时", "公式"], ["赵六", 5, None], ["孙七", 4, None], ["周八", 3, None], ["吴九", 2, "=B5*A1"]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"),),
    )

    summary = MergeEngine().execute(job)

    workbook = load_workbook(output)
    # 锁定当前平移行为：=B5*A1 从第 5 行追加到第 8 行 → =B8*A4（Excel 复制语义，可能指错位置）
    assert workbook["工时"]["C8"].value == "=B8*A4"
    workbook.close()
    warnings = summary.results[0].warnings
    assert any("A1" in w and "绝对引用" in w and "部门B.xlsx" in w for w in warnings)


def test_formula_absolute_and_cross_sheet_references_do_not_warn(tmp_path):
    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {"工时": [["姓名", "工时", "公式"], ["张三", 8, None]], "汇总": [["指标"], [1]]},
    )
    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {
            "工时": [
                ["姓名", "工时", "公式"],
                ["李四", 7, "=$A$1*B2"],      # 绝对引用：平移安全，不警告
                ["王五", 6, "=汇总!A1*B3"],   # 跨 sheet 引用：不在本警告范围
                ["赵六", 5, "=B4+C4"],        # 引用同区域数据行：正常平移，不警告
            ],
            "汇总": [["指标"], [2]],
        },
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(MergeSheetConfig("工时"), MergeSheetConfig("汇总")),
    )

    summary = MergeEngine().execute(job)

    result = next(r for r in summary.results if r.sheet_name == "工时")
    assert not any("表头行上方" in w for w in result.warnings)
    workbook = load_workbook(output)
    sheet = workbook["工时"]
    assert sheet["C3"].value == "=$A$1*B3"
    assert sheet["C4"].value == "=汇总!A2*B4"  # 跨 sheet 相对引用也会被平移（已锁定语义）
    assert sheet["C5"].value == "=B5+C5"
    workbook.close()


def test_merge_extends_chart_data_ranges(tmp_path):
    from openpyxl.chart import BarChart, Reference

    file_a = _make_workbook(
        tmp_path / "部门A.xlsx",
        {
            "数据": [["类别", "值"], ["甲", 1], ["乙", 2], ["丙", 3], ["丁", 4]],
            "汇总": [["指标"], [100]],
        },
    )
    workbook = load_workbook(file_a)
    data_ws = workbook["数据"]
    # 图表放在被合并的「数据」上，引用「数据」的 A2:A5 / B2:B5
    chart_on_data = BarChart()
    chart_on_data.add_data(Reference(data_ws, min_col=2, min_row=2, max_row=5), titles_from_data=False)
    chart_on_data.set_categories(Reference(data_ws, min_col=1, min_row=2, max_row=5))
    data_ws.add_chart(chart_on_data, "E2")
    # 图表放在未被合并的「汇总」上，但引用被合并的「数据」
    chart_on_summary = BarChart()
    chart_on_summary.add_data(Reference(data_ws, min_col=2, min_row=2, max_row=5), titles_from_data=False)
    workbook["汇总"].add_chart(chart_on_summary, "C2")
    # 引用「汇总」自身（该 sheet 标记 identical 不追加）的图表不应变化
    chart_self = BarChart()
    chart_self.add_data(Reference(workbook["汇总"], min_col=1, min_row=2, max_row=2), titles_from_data=False)
    workbook["汇总"].add_chart(chart_self, "C20")
    workbook.save(file_a)
    workbook.close()

    file_b = _make_workbook(
        tmp_path / "部门B.xlsx",
        {"数据": [["类别", "值"], ["戊", 5], ["己", 6], ["庚", 7]]},
    )
    output = tmp_path / "合并结果.xlsx"
    job = MergeJob(
        input_files=(file_a, file_b),
        output_file=output,
        sheet_configs=(
            MergeSheetConfig("数据"),
            MergeSheetConfig("汇总", identical=True),
        ),
    )

    MergeEngine().execute(job)

    workbook = load_workbook(output)
    data_charts = workbook["数据"]._charts
    assert len(data_charts) == 1
    series = data_charts[0].series[0]
    # 追加 3 行后数据末行从 5 延伸到 8
    assert series.val.numRef.f == "'数据'!$B$2:$B$8"
    assert series.cat.numRef.f == "'数据'!$A$2:$A$8"
    summary_charts = workbook["汇总"]._charts
    assert len(summary_charts) == 2
    # 未合并 sheet 上的图表，引用被合并 sheet 的也被延伸
    assert summary_charts[0].series[0].val.numRef.f == "'数据'!$B$2:$B$8"
    # identical sheet 未被追加，引用自身的图表不动（openpyxl 会把单单元格区域折叠成单格引用）
    assert summary_charts[1].series[0].val.numRef.f == "'汇总'!$A$2"
    workbook.close()
