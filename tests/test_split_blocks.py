from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from excel_splitter.engine import SplitEngine
from excel_splitter.models import SheetConfig, SplitJob
from excel_splitter.planning import build_split_plan
from excel_splitter.web.app import create_app


def _block_workbook(tmp_path: Path) -> Path:
    """仿真实结构：主表 + 合计行 + 两个小表（空行分隔）。"""
    path = tmp_path / "多表区.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "概览"
    sheet.append(["年度概览", None])       # r1 标题
    sheet.append(["团队", "指标"])          # r2 主表头
    sheet.append(["甲", 1])
    sheet.append(["乙", 2])
    sheet.append(["丙", 3])                # r5
    sheet.append([None, 100])              # r6 合计行
    sheet.append([None, None])             # r7 空行
    sheet.append(["负责人", "人数"])        # r8 小表1表头
    sheet.append(["甲", 10])
    sheet.append(["甲", 20])               # r10
    sheet.append([None, None])             # r11 空行
    sheet.append(["负责人", "项目"])        # r12 小表2表头
    sheet.append(["甲", "P1"])
    sheet.append(["乙", "P2"])
    sheet.append(["乙", "P3"])             # r15
    for row in (8, 12):
        for col in (1, 2):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
    workbook.save(path)
    workbook.close()
    return path


def _job(input_file: Path, output_dir: Path, value: str, strategies=()) -> SplitJob:
    return SplitJob(
        input_file=input_file,
        output_dir=output_dir,
        sheet_configs=(
            SheetConfig("概览", 2, 1, "A - 团队", block_strategies=tuple(strategies)),
        ),
        split_mode="selected",
        selected_split_values=(value,),
        output_types=("formula",),
    )


def _formula_rows(result, sheet_name="概览"):
    output_file = next(
        item.output_file for item in result.output_files if item.output_type == "formula"
    )
    workbook = load_workbook(output_file)
    rows = [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
    workbook.close()
    return rows


def test_plan_detects_single_block_sheet(tmp_path):
    path = tmp_path / "单表.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(["团队", "指标"])
    sheet.append(["甲", 1])
    sheet.append(["乙", 2])
    workbook.save(path)
    workbook.close()

    plan = build_split_plan(path, (SheetConfig("数据", 1, 1),))

    assert plan.blocks["数据"] == [(1, 2, 3)]


def test_plan_detects_multi_blocks_with_gap_and_total_row(tmp_path):
    plan = build_split_plan(
        _block_workbook(tmp_path), (SheetConfig("概览", 2, 1),)
    )

    assert plan.blocks["概览"] == [(2, 3, 6), (8, 9, 10), (12, 13, 15)]


def test_follow_split_keeps_block_headers_and_gap_rows(tmp_path):
    summary = SplitEngine().execute(
        _job(_block_workbook(tmp_path), tmp_path / "输出", "甲")
    )

    result = summary.results[0]
    assert result.split_value == "甲"
    rows = _formula_rows(result)
    # 标题 + 主表头 + 甲 + 空行 + 小表1表头 + 2 行 + 空行 + 小表2表头 + 1 行
    assert [row[0] for row in rows] == [
        "年度概览", "团队", "甲", None, "负责人", "甲", "甲", None, "负责人", "甲",
    ]
    assert rows[4] == ["负责人", "人数"]
    assert rows[8] == ["负责人", "项目"]
    # 合计行被丢弃
    assert all(row[1] != 100 for row in rows)
    assert result.discarded_empty_rows["概览"] == 1  # 合计行（键为空）

    workbook = load_workbook(
        next(item.output_file for item in result.output_files if item.output_type == "formula")
    )
    # 块表头样式随保留行原样存在
    assert workbook["概览"]["A5"].font.bold
    assert workbook["概览"]["A9"].font.bold
    assert workbook["概览"]["A5"].fill.fgColor.rgb.endswith("D9EAF7")
    workbook.close()


def test_follow_split_drops_unmatched_block_entirely(tmp_path):
    summary = SplitEngine().execute(
        _job(_block_workbook(tmp_path), tmp_path / "输出", "乙")
    )

    rows = _formula_rows(summary.results[0])
    # 小表1 无匹配行：整块（含表头）消失，其周边空行也一并删除
    assert [row[0] for row in rows] == [
        "年度概览", "团队", "乙", "负责人", "乙", "乙",
    ]
    assert rows[3] == ["负责人", "项目"]


def test_follow_split_without_any_matching_block_keeps_main_only(tmp_path):
    summary = SplitEngine().execute(
        _job(_block_workbook(tmp_path), tmp_path / "输出", "丙")
    )

    rows = _formula_rows(summary.results[0])
    assert [row[0] for row in rows] == ["年度概览", "团队", "丙"]


def test_keep_strategy_retains_block_in_every_output(tmp_path):
    summary = SplitEngine().execute(
        _job(_block_workbook(tmp_path), tmp_path / "输出", "乙", strategies=("keep",))
    )

    rows = _formula_rows(summary.results[0])
    # 乙 在小表1 无匹配行，但 keep 策略让整块原样保留
    assert [row[0] for row in rows] == [
        "年度概览", "团队", "乙", None, "负责人", "甲", "甲", None, "负责人", "乙", "乙",
    ]
    assert rows[4] == ["负责人", "人数"]


def test_drop_strategy_removes_block_everywhere(tmp_path):
    summary = SplitEngine().execute(
        _job(_block_workbook(tmp_path), tmp_path / "输出", "甲", strategies=("follow", "drop"))
    )

    rows = _formula_rows(summary.results[0])
    # 小表2 被 drop：即使甲有匹配行也整块消失
    assert [row[0] for row in rows] == [
        "年度概览", "团队", "甲", None, "负责人", "甲", "甲",
    ]


def test_linked_mode_applies_block_logic(tmp_path):
    path = tmp_path / "关联.xlsx"
    workbook = Workbook()
    reference = workbook.active
    reference.title = "归属"
    reference.append(["团队", "姓名"])
    reference.append(["甲", "张三"])
    reference.append(["乙", "李四"])
    detail = workbook.create_sheet("明细")
    detail.append(["事项", "姓名"])
    detail.append(["主表项1", "张三"])
    detail.append(["主表项2", "李四"])
    detail.append([None, None])
    detail.append(["小表事项", "姓名"])
    detail.append(["小表项1", "张三"])
    detail.append(["小表项2", "王五"])
    workbook.save(path)
    workbook.close()

    job = SplitJob(
        input_file=path,
        output_dir=tmp_path / "输出",
        sheet_configs=(
            SheetConfig("归属", 1, 1, "A - 团队", mode="reference", key_column_idx=2),
            SheetConfig("明细", 1, None, mode="linked", key_column_idx=2),
        ),
        split_mode="selected",
        selected_split_values=("甲",),
        output_types=("formula",),
    )
    summary = SplitEngine().execute(job)

    result = summary.results[0]
    rows = _formula_rows(result, sheet_name="明细")
    assert [row[0] for row in rows] == ["事项", "主表项1", None, "小表事项", "小表项1"]
    assert result.unmatched_key_rows["明细"] == 1  # 王五不在基准键集中


def _app(tmp_path):
    return create_app(
        {
            "TESTING": True,
            "UPLOAD_DIR": tmp_path / "uploads",
            "DEFAULT_OUTPUT_DIR": tmp_path / "outputs",
        }
    )


def test_routes_return_block_info_and_pass_strategies(tmp_path):
    from io import BytesIO

    client = _app(tmp_path).test_client()
    source = _block_workbook(tmp_path)
    with source.open("rb") as handle:
        loaded = client.post(
            "/api/load",
            data={"file": (BytesIO(handle.read()), "多表区.xlsx")},
            content_type="multipart/form-data",
        )
    job_id = loaded.get_json()["job_id"]
    config = {
        "sheet_name": "概览",
        "header_row": 2,
        "split_column_idx": 1,
        "split_column_label": "A - 团队",
        "mode": "direct",
    }

    values = client.post(
        "/api/split-values", json={"job_id": job_id, "sheet_configs": [config]}
    )
    assert values.status_code == 200
    blocks = values.get_json()["blocks"]["概览"]
    assert [(b["header_row"], b["data_start"], b["data_end"]) for b in blocks] == [
        (2, 3, 6),
        (8, 9, 10),
        (12, 13, 15),
    ]
    assert blocks[1]["header_preview"] == "负责人 / 人数"

    executed = client.post(
        "/api/execute",
        json={
            "job_id": job_id,
            "sheet_configs": [{**config, "block_strategies": ["follow", "drop"]}],
            "split_mode": "selected",
            "selected_split_values": ["甲"],
            "output_types": ["formula"],
            "output_dir": str(tmp_path / "输出"),
        },
    )
    assert executed.status_code == 200
    output_file = Path(
        executed.get_json()["results"][0]["output_files"][0]["output_file"]
    )
    workbook = load_workbook(output_file)
    rows = [row[0] for row in workbook["概览"].iter_rows(values_only=True)]
    workbook.close()
    # drop 策略透传生效：小表2 整块消失
    assert rows == ["年度概览", "团队", "甲", None, "负责人", "甲", "甲"]
