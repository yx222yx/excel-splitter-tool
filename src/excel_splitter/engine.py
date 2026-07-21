from __future__ import annotations

import os
from copy import copy
from concurrent.futures import ProcessPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Callable, Iterable

from .excel_io import load_workbook_with_warnings
from .file_utils import render_filename, resolve_output_path
from .models import OutputArtifact, SheetConfig, SplitJob, SplitResult, SplitSummary
from .planning import SplitPlan, build_split_plan
from .values import normalize_split_value


EMPTY_ROW_STOP_THRESHOLD = 10000
ProgressCallback = Callable[[int, str], None]
DEFAULT_PROCESS_WORKERS = 2
MAX_PROCESS_WORKERS = 3


def _select_process_workers(
    value_count: int,
    requested: int | None,
) -> int:
    if value_count <= 1:
        return 1
    if requested is None:
        raw = os.environ.get("EXCEL_SPLITTER_WORKERS", "").strip()
        try:
            requested = int(raw) if raw else DEFAULT_PROCESS_WORKERS
        except ValueError:
            requested = DEFAULT_PROCESS_WORKERS
    available_cpus = max(1, os.cpu_count() or 1)
    return max(
        1,
        min(value_count, requested, MAX_PROCESS_WORKERS, available_cpus),
    )


def _export_value_worker(
    job: SplitJob,
    split_value: str,
    plan: SplitPlan,
    file_bytes: bytes,
) -> SplitResult:
    engine = SplitEngine(process_workers=1)
    return engine._export_value(
        job,
        split_value,
        plan,
        lambda _fraction, _message: None,
        file_bytes=file_bytes,
    )


class SplitEngine:
    def __init__(self, process_workers: int | None = None) -> None:
        self.process_workers = process_workers

    def execute(
        self,
        job: SplitJob,
        progress_callback: ProgressCallback | None = None,
        plan: SplitPlan | None = None,
    ) -> SplitSummary:
        progress = _ProgressReporter(progress_callback)
        progress.update(0, "正在分析工作簿")
        job.validate()

        def report_plan_progress(percent: int, message: str) -> None:
            progress.update(min(4, round(percent * 4 / 100)), message)

        if plan is None:
            plan = build_split_plan(
                job.input_file, job.sheet_configs,
                progress_callback=report_plan_progress,
            )
        target_values = self._target_values(job, plan.values)
        if not target_values:
            raise ValueError("没有可用于拆分的非空值")
        progress.update(5, f"已识别 {len(target_values)} 个拆分值")

        job.output_dir.mkdir(parents=True, exist_ok=True)
        n = len(target_values)
        errors: list[str] = []
        all_warnings = list(plan.warnings)
        lock = Lock()

        # 文件只读一次，传给所有并行 worker 共享
        with open(job.input_file, "rb") as f:
            file_bytes = f.read()

        # 预分配结果列表，按顺序填入
        results: list[SplitResult | None] = [None] * n
        worker_progress = [0.0] * n

        def report_worker_progress(idx: int, fraction: float, message: str) -> None:
            clipped = max(0.0, min(1.0, fraction))
            with lock:
                worker_progress[idx] = max(worker_progress[idx], clipped)
                pct = 5 + int(sum(worker_progress) * 90 / n)
                progress.update(pct, f"{message} ({idx + 1}/{n})")

        def mark_worker_done(idx: int) -> None:
            with lock:
                worker_progress[idx] = 1.0
                completed = sum(1 for item in worker_progress if item >= 1.0)
                pct = 5 + int(sum(worker_progress) * 90 / n)
                message = f"正在拆分 ({completed}/{n})"
                progress.update(pct, message)

        def process_one(idx: int, sv: str) -> None:
            try:
                r = self._export_value(
                    job,
                    sv,
                    plan,
                    lambda fraction, message: report_worker_progress(
                        idx, fraction, message
                    ),
                    file_bytes=file_bytes,
                )
                with lock:
                    results[idx] = r
                    all_warnings.extend(r.warnings)
            except Exception as exc:
                with lock:
                    errors.append(f"{sv}: {exc}")
            finally:
                mark_worker_done(idx)

        workers = _select_process_workers(n, self.process_workers)
        use_processes = workers > 1 and type(self) is SplitEngine
        if use_processes:
            with ProcessPoolExecutor(max_workers=workers) as pool:
                future_jobs = {}
                for idx, split_value in enumerate(target_values):
                    report_worker_progress(idx, 0.02, f"正在加载 {split_value}")
                    future = pool.submit(
                        _export_value_worker,
                        job,
                        split_value,
                        plan,
                        file_bytes,
                    )
                    future_jobs[future] = (idx, split_value)
                for future in as_completed(future_jobs):
                    idx, split_value = future_jobs[future]
                    try:
                        result = future.result()
                        results[idx] = result
                        all_warnings.extend(result.warnings)
                    except Exception as exc:
                        errors.append(f"{split_value}: {exc}")
                    finally:
                        mark_worker_done(idx)
        else:
            for idx, split_value in enumerate(target_values):
                process_one(idx, split_value)

        ordered = [r for r in results if r is not None]
        unique_warnings = tuple(dict.fromkeys(all_warnings))
        summary = SplitSummary(
            results=ordered,
            total_files=sum(len(r.output_files) for r in ordered),
            total_discarded=sum(
                sum(r.discarded_empty_rows.values()) for r in ordered
            ),
            total_unmatched=sum(
                sum(r.unmatched_key_rows.values()) for r in ordered
            ),
            warnings=list(unique_warnings),
            errors=errors,
        )
        progress.update(100, "拆分完成")
        return summary

    def _target_values(self, job: SplitJob, available_values: list[str]) -> list[str]:
        if job.split_mode == "all":
            return available_values
        selected = _unique_normalized(job.selected_split_values)
        unavailable = [value for value in selected if value not in available_values]
        if unavailable:
            raise ValueError(f"所选拆分值不存在：{', '.join(unavailable)}")
        return selected

    def _export_value(
        self,
        job: SplitJob,
        split_value: str,
        plan: SplitPlan,
        progress: Callable[[float, str], None],
        *,
        file_bytes: bytes | None = None,
    ) -> SplitResult:
        progress(0.02, f"正在加载 {split_value}")
        formula_workbook = None
        value_workbook = None
        warning_messages: list[str] = []
        need_both = "formula" in job.output_types
        try:
            if file_bytes is None:
                with open(job.input_file, "rb") as f:
                    file_bytes = f.read()
            if need_both:
                formula_workbook, formula_warnings = load_workbook_with_warnings(
                    BytesIO(file_bytes), data_only=False
                )
                warning_messages.extend(formula_warnings)
            value_workbook, value_warnings = load_workbook_with_warnings(
                BytesIO(file_bytes), data_only=True
            )
            warning_messages = list(
                dict.fromkeys([*warning_messages, *value_warnings])
            )
            selected = set(job.selected_sheets)
            if formula_workbook is not None:
                for sheet_name in list(formula_workbook.sheetnames):
                    if sheet_name not in selected:
                        formula_workbook.remove(formula_workbook[sheet_name])
            for sheet_name in list(value_workbook.sheetnames):
                if sheet_name not in selected:
                    value_workbook.remove(value_workbook[sheet_name])

            sheet_rows: dict[str, int] = {}
            discarded_empty_rows: dict[str, int] = {}
            unmatched_key_rows: dict[str, int] = {}
            kept_rows_per_sheet: dict[str, list[int]] = {}
            config_count = len(job.sheet_configs)
            for config_index, config in enumerate(job.sheet_configs):
                sheet = (
                    formula_workbook[config.sheet_name]
                    if formula_workbook is not None
                    else None
                )
                value_sheet = value_workbook[config.sheet_name]
                filter_start = 0.12 + (config_index * 0.58 / config_count)
                filter_end = 0.12 + ((config_index + 1) * 0.58 / config_count)

                def report_rows(current: int, total: int) -> None:
                    ratio = current / total if total else 1
                    progress(
                        filter_start + (filter_end - filter_start) * ratio,
                        f"正在筛选 {split_value} / {config.sheet_name} ({current}/{total})",
                    )

                if config.mode == "full":
                    kept = max(0, value_sheet.max_row - config.header_row)
                    discarded = 0
                    unmatched = 0
                    report_rows(kept, kept)
                elif config.mode == "linked":
                    kept, discarded, unmatched, kept_original = _filter_linked_sheet(
                        sheet,
                        value_sheet,
                        config,
                        plan.keys_by_value.get(split_value, set()),
                        plan.all_keys,
                        report_rows,
                        blocks=plan.blocks.get(config.sheet_name),
                    )
                    kept_rows_per_sheet[config.sheet_name] = kept_original
                else:
                    kept, discarded, kept_original = _filter_sheet(
                        sheet,
                        value_sheet,
                        config,
                        split_value,
                        report_rows,
                        blocks=plan.blocks.get(config.sheet_name),
                    )
                    unmatched = 0
                if config.mode != "full":
                    _compact_filtered_sheet(value_sheet, config, kept_original)
                    if sheet is not None:
                        _compact_filtered_sheet(sheet, config, kept_original)
                    kept_rows_per_sheet[config.sheet_name] = kept_original
                sheet_rows[config.sheet_name] = kept
                discarded_empty_rows[config.sheet_name] = discarded
                unmatched_key_rows[config.sheet_name] = unmatched

            if formula_workbook is not None:
                _fix_formula_references(
                    formula_workbook, job.sheet_configs, kept_rows_per_sheet
                )

            available_outputs = {
                "formula": ("公式版", formula_workbook),
                "values": ("结果值版", value_workbook),
            }
            output_files: list[OutputArtifact] = []
            for output_index, output_type in enumerate(job.output_types):
                output_label, output_workbook = available_outputs[output_type]
                progress(
                    0.75 + output_index * 0.23 / len(job.output_types),
                    f"正在保存 {split_value} / {output_label}",
                )
                filename = render_filename(
                    job.filename_template,
                    original_name=job.original_name or job.input_file.stem,
                    split_value=split_value,
                    output_type=output_label,
                )
                output_file = resolve_output_path(
                    job.output_dir / filename, overwrite=job.overwrite
                )
                output_workbook.save(output_file)
                output_files.append(
                    OutputArtifact(
                        output_type=output_type,
                        output_file=output_file,
                    )
                )
                progress(
                    0.75 + (output_index + 1) * 0.23 / len(job.output_types),
                    f"已保存 {split_value} / {output_label}",
                )
        finally:
            if formula_workbook is not None:
                formula_workbook.close()
            if value_workbook is not None:
                value_workbook.close()

        return SplitResult(
            split_value=split_value,
            output_files=output_files,
            sheet_rows=sheet_rows,
            discarded_empty_rows=discarded_empty_rows,
            unmatched_key_rows=unmatched_key_rows,
            warnings=warning_messages,
        )


def _effective_last_data_row(value_sheet, header_row: int) -> int:
    cells = getattr(value_sheet, "_cells", None)
    if isinstance(cells, dict):
        last_data_row = header_row
        for key, cell in cells.items():
            row_index = key[0]
            if row_index > header_row and normalize_split_value(cell.value) is not None:
                last_data_row = max(last_data_row, row_index)
        return last_data_row

    max_column = getattr(value_sheet, "max_column", 1)
    last_data_row = header_row
    consecutive_empty = 0
    for row_index in range(header_row + 1, value_sheet.max_row + 1):
        has_value = any(
            normalize_split_value(value_sheet.cell(row=row_index, column=column).value)
            is not None
            for column in range(1, max_column + 1)
        )
        if has_value:
            last_data_row = row_index
            consecutive_empty = 0
        else:
            consecutive_empty += 1
            if consecutive_empty >= EMPTY_ROW_STOP_THRESHOLD:
                break
    return last_data_row


def _add_delete_range(
    ranges: list[tuple[int, int]],
    pending: tuple[int, int] | None,
    row_index: int,
) -> tuple[int, int]:
    if pending is None:
        return (row_index, row_index)
    start, end = pending
    if row_index == start - 1:
        return (row_index, end)
    ranges.append(pending)
    return (row_index, row_index)


def _flush_delete_range(
    ranges: list[tuple[int, int]], pending: tuple[int, int] | None
) -> None:
    if pending is not None:
        ranges.append(pending)


def _delete_rows_both(sheet, value_sheet, start: int, amount: int) -> None:
    if amount <= 0:
        return
    if sheet is not None and start <= sheet.max_row:
        sheet.delete_rows(start, min(amount, sheet.max_row - start + 1))
    if start <= value_sheet.max_row:
        value_sheet.delete_rows(start, min(amount, value_sheet.max_row - start + 1))


def _delete_row_ranges(sheet, value_sheet, ranges: list[tuple[int, int]]) -> None:
    for start, end in ranges:
        _delete_rows_both(sheet, value_sheet, start, end - start + 1)


def _filter_sheet(
    sheet,
    value_sheet,
    config: SheetConfig,
    split_value: str,
    progress: Callable[[int, int], None] | None = None,
    blocks: list[tuple[int, int, int]] | None = None,
) -> tuple[int, int, list[int]]:
    if blocks and len(blocks) > 1:
        kept, discarded_empty, _unmatched, kept_original_rows = _filter_blocks(
            sheet,
            value_sheet,
            config,
            blocks,
            key_column_idx=config.split_column_idx,
            matches=lambda key: key == split_value,
            unmatched_check=None,
            progress=progress,
        )
        return kept, discarded_empty, kept_original_rows
    kept = 0
    discarded_empty = 0
    kept_original_rows: list[int] = []
    assert config.split_column_idx is not None
    original_max_row = value_sheet.max_row
    effective_last_row = _effective_last_data_row(value_sheet, config.header_row)
    total = max(0, effective_last_row - config.header_row)
    delete_ranges: list[tuple[int, int]] = []
    pending_delete: tuple[int, int] | None = None
    if effective_last_row < original_max_row:
        delete_ranges.append((effective_last_row + 1, original_max_row))
    for processed, row_index in enumerate(
        range(effective_last_row, config.header_row, -1), start=1
    ):
        normalized = normalize_split_value(
            value_sheet.cell(row=row_index, column=config.split_column_idx).value
        )
        if normalized is None:
            discarded_empty += 1
            pending_delete = _add_delete_range(
                delete_ranges, pending_delete, row_index
            )
        elif normalized != split_value:
            pending_delete = _add_delete_range(
                delete_ranges, pending_delete, row_index
            )
        else:
            _flush_delete_range(delete_ranges, pending_delete)
            pending_delete = None
            kept += 1
            kept_original_rows.append(row_index)
        if progress is not None and (processed % 25 == 0 or processed == total):
            progress(processed, total)
    _flush_delete_range(delete_ranges, pending_delete)
    _delete_row_ranges(sheet, value_sheet, delete_ranges)
    return kept, discarded_empty, kept_original_rows


def _filter_linked_sheet(
    sheet,
    value_sheet,
    config: SheetConfig,
    target_keys: set[str],
    all_keys: set[str],
    progress: Callable[[int, int], None] | None = None,
    blocks: list[tuple[int, int, int]] | None = None,
) -> tuple[int, int, int, list[int]]:
    if blocks and len(blocks) > 1:
        return _filter_blocks(
            sheet,
            value_sheet,
            config,
            blocks,
            key_column_idx=config.key_column_idx,
            matches=lambda key: key in target_keys,
            unmatched_check=lambda key: key not in all_keys,
            progress=progress,
        )
    kept = 0
    discarded_empty = 0
    unmatched = 0
    kept_original_rows: list[int] = []
    assert config.key_column_idx is not None
    original_max_row = value_sheet.max_row
    effective_last_row = _effective_last_data_row(value_sheet, config.header_row)
    total = max(0, effective_last_row - config.header_row)
    delete_ranges: list[tuple[int, int]] = []
    pending_delete: tuple[int, int] | None = None
    if effective_last_row < original_max_row:
        delete_ranges.append((effective_last_row + 1, original_max_row))
    for processed, row_index in enumerate(
        range(effective_last_row, config.header_row, -1), start=1
    ):
        key = normalize_split_value(
            value_sheet.cell(row=row_index, column=config.key_column_idx).value
        )
        if key is None:
            discarded_empty += 1
            pending_delete = _add_delete_range(
                delete_ranges, pending_delete, row_index
            )
        elif key in target_keys:
            _flush_delete_range(delete_ranges, pending_delete)
            pending_delete = None
            kept += 1
            kept_original_rows.append(row_index)
        else:
            if key not in all_keys:
                unmatched += 1
            pending_delete = _add_delete_range(
                delete_ranges, pending_delete, row_index
            )
        if progress is not None and (processed % 25 == 0 or processed == total):
            progress(processed, total)
    _flush_delete_range(delete_ranges, pending_delete)
    _delete_row_ranges(sheet, value_sheet, delete_ranges)
    return kept, discarded_empty, unmatched, kept_original_rows


def _filter_blocks(
    sheet,
    value_sheet,
    config: SheetConfig,
    blocks: list[tuple[int, int, int]],
    *,
    key_column_idx: int | None,
    matches: Callable[[str], bool],
    unmatched_check: Callable[[str], bool] | None,
    progress: Callable[[int, int], None] | None = None,
) -> tuple[int, int, int, list[int]]:
    """多表区 sheet 的按块过滤。

    块 1（主表）行为与单行过滤一致；块 2..n 按策略处理：follow 按同样的键
    匹配，块内有幸存数据行才保留块表头；keep 整块保留；drop 整块删除（与
    不匹配行一样不计入统计）。幸存块之间保留一个空行作为分隔。
    """
    assert key_column_idx is not None
    strategies = config.block_strategies
    original_max_row = value_sheet.max_row
    effective_last_row = _effective_last_data_row(value_sheet, config.header_row)
    total = max(0, effective_last_row - config.header_row)
    delete_rows: set[int] = set()
    if effective_last_row < original_max_row:
        delete_rows.update(range(effective_last_row + 1, original_max_row + 1))
    kept = 0
    discarded_empty = 0
    unmatched = 0
    kept_original_rows: list[int] = []
    survived: list[bool] = []
    processed = 0
    for block_index, (block_header, data_start, data_end) in enumerate(blocks):
        strategy = (
            "follow"
            if block_index == 0
            else strategies[block_index - 1]
            if block_index - 1 < len(strategies)
            else "follow"
        )
        if strategy == "keep":
            if block_index > 0:
                kept_original_rows.append(block_header)
            for row_index in range(data_start, data_end + 1):
                kept += 1
                kept_original_rows.append(row_index)
                processed += 1
            survived.append(True)
            continue
        if strategy == "drop":
            if block_index > 0:
                delete_rows.add(block_header)
            delete_rows.update(range(data_start, data_end + 1))
            processed += max(0, data_end - data_start + 1)
            survived.append(False)
            continue
        block_kept = 0
        for row_index in range(data_start, data_end + 1):
            key = normalize_split_value(
                value_sheet.cell(row=row_index, column=key_column_idx).value
            )
            if key is None:
                discarded_empty += 1
                delete_rows.add(row_index)
            elif matches(key):
                kept += 1
                block_kept += 1
                kept_original_rows.append(row_index)
            else:
                if unmatched_check is not None and unmatched_check(key):
                    unmatched += 1
                delete_rows.add(row_index)
            processed += 1
            if progress is not None and (processed % 25 == 0 or processed == total):
                progress(processed, total)
        if block_index == 0:
            survived.append(True)
        elif block_kept:
            kept_original_rows.append(block_header)
            survived.append(True)
        else:
            delete_rows.add(block_header)
            survived.append(False)
    # 空行分隔：相邻两个幸存块之间保留一个空行，其余空行删除
    for block_index in range(1, len(blocks)):
        gap_start = blocks[block_index - 1][2] + 1
        gap_end = blocks[block_index][0] - 1
        if gap_start > gap_end:
            continue
        if survived[block_index - 1] and survived[block_index]:
            kept_original_rows.append(gap_start)
            delete_rows.update(range(gap_start + 1, gap_end + 1))
        else:
            delete_rows.update(range(gap_start, gap_end + 1))
    kept_original_rows.sort()
    _delete_row_ranges(sheet, value_sheet, _rows_to_ranges(delete_rows))
    return kept, discarded_empty, unmatched, kept_original_rows


def _rows_to_ranges(rows) -> list[tuple[int, int]]:
    """把行号集合压缩成自下而上排序的连续区间，供从底向上删行。"""
    ranges: list[tuple[int, int]] = []
    for row_index in sorted(rows, reverse=True):
        if ranges and ranges[-1][0] == row_index + 1:
            ranges[-1] = (row_index, ranges[-1][1])
        else:
            ranges.append((row_index, row_index))
    return ranges
def _compact_filtered_sheet(
    sheet,
    config: SheetConfig,
    kept_original_rows: Iterable[int],
) -> None:
    from openpyxl.utils.cell import get_column_letter, range_boundaries
    from openpyxl.worksheet.dimensions import DimensionHolder, RowDimension

    kept_rows = sorted(kept_original_rows)
    row_mapping = [
        *((row_index, row_index) for row_index in range(1, config.header_row + 1)),
        *(
            (original_row, config.header_row + offset)
            for offset, original_row in enumerate(kept_rows, start=1)
        ),
    ]
    source_dimensions = sheet.row_dimensions
    compact_dimensions = DimensionHolder(
        worksheet=sheet,
        default_factory=lambda: RowDimension(sheet),
    )
    for original_row, compact_row in row_mapping:
        if original_row not in source_dimensions:
            continue
        dimension = copy(source_dimensions[original_row])
        dimension.index = compact_row
        compact_dimensions[compact_row] = dimension
    sheet.row_dimensions = compact_dimensions

    last_row = max(config.header_row, sheet.max_row)

    def trim_ref(ref: str | None) -> str | None:
        if not ref:
            return ref
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        if max_row <= last_row or min_row > last_row:
            return ref
        return (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{last_row}"
        )

    sheet.auto_filter.ref = trim_ref(sheet.auto_filter.ref)
    for table in sheet.tables.values():
        table.ref = trim_ref(table.ref)
        if table.autoFilter is not None:
            table.autoFilter.ref = trim_ref(table.autoFilter.ref)




def _fix_formula_references(
    workbook,
    sheet_configs: tuple[SheetConfig, ...],
    kept_rows_per_sheet: dict[str, list[int]],
) -> None:
    """修正 delete_rows 后偏移行的公式引用。

    从底向上删除行时，被保留的数据行可能因上方行的删除而向上移位，
    但 openpyxl 的 delete_rows 不会自动重写公式文本。
    此函数逐 cell 计算偏移量并调用 Translator 修正公式中的引用。
    """
    from openpyxl.formula.translate import Translator

    for config in sheet_configs:
        sheet_name = config.sheet_name
        kept_original = kept_rows_per_sheet.get(sheet_name)
        if not kept_original:
            continue
        kept_original.sort()
        ws = workbook[sheet_name]
        original_by_compact_row = {
            config.header_row + data_index: original_row
            for data_index, original_row in enumerate(kept_original, start=1)
        }
        cells = getattr(ws, "_cells", {})
        for (row_index, _column_index), cell in cells.items():
            original_row = original_by_compact_row.get(row_index)
            if original_row is None or original_row == row_index:
                continue
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                continue
            offset = row_index - original_row
            try:
                cell.value = Translator(cell.value, cell.coordinate).translate_formula(
                    row_delta=offset, col_delta=0
                )
            except Exception:
                pass



class _ProgressReporter:
    def __init__(self, callback: ProgressCallback | None) -> None:
        self.callback = callback
        self.last_percent = -1

    def update(self, percent: int, message: str) -> None:
        normalized = max(self.last_percent, min(100, max(0, percent)))
        if self.callback is not None and (
            normalized != self.last_percent or normalized in (0, 100)
        ):
            self.callback(normalized, message)
        self.last_percent = normalized


def _unique_normalized(values: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = normalize_split_value(value)
        if normalized is not None and normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    return result
