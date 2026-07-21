from __future__ import annotations

import hashlib
import re
from copy import copy
from pathlib import Path
from typing import Callable
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from .excel_io import load_workbook_with_warnings
from .merge_models import MergeJob, MergeSheetConfig, MergeSheetResult, MergeSummary
from .merge_planning import (
    MergePlan,
    MergeSheetPlan,
    _identities,
    build_merge_plan,
)
from .values import normalize_split_value


EMPTY_ROW_STOP_THRESHOLD = 10000
EXTRA_COLUMN_WIDTH = 12  # 并集新增列的默认列宽
ProgressCallback = Callable[[int, str], None]

# 常见 A1 形态单元格引用（排除跨 sheet 引用、函数名、区域右半部分）
_CELL_REF_PATTERN = re.compile(r"(?<![A-Za-z0-9_$!:])(\$?)([A-Z]{1,3})(\$?)([0-9]{1,7})(?!\()")


class MergeEngine:
    """模板照抄式合并引擎。

    以第一个输入文件为输出模板（普通模式加载，公式/样式/合并/冻结等全部
    保留），只删除未选中的 sheet；选中 sheet 的表头行及之前内容一个字节不动，
    其余文件的数据行按表头名对齐后流式追加到末尾。
    """

    def execute(
        self,
        job: MergeJob,
        progress_callback: ProgressCallback | None = None,
        plan: MergePlan | None = None,
    ) -> MergeSummary:
        progress = _ProgressReporter(progress_callback)
        progress.update(0, "正在分析输入文件")
        job.validate()

        # identical 的 sheet 不需要字段扫描，其余 sheet 扫描表头生成并集与映射
        scan_configs = tuple(c for c in job.sheet_configs if not c.identical)
        if plan is None:
            plan = (
                build_merge_plan(job.input_files, scan_configs)
                if scan_configs
                else MergePlan(sheets=[])
            )
        plans_by_sheet = {sheet.sheet_name: sheet for sheet in plan.sheets}
        progress.update(5, "正在加载模板工作簿")

        first_file = job.input_files[0]
        workbook, load_warnings = load_workbook_with_warnings(first_file, data_only=False)
        all_warnings = [*plan.warnings, *load_warnings]
        selected = set(job.selected_sheets)
        for name in list(workbook.sheetnames):
            if name not in selected:
                workbook.remove(workbook[name])

        errors: list[str] = []
        source_rows: dict[str, dict[str, int]] = {
            config.sheet_name: {} for config in job.sheet_configs
        }
        skipped_duplicates: dict[str, dict[str, str]] = {
            config.sheet_name: {} for config in job.sheet_configs
        }
        fingerprints: dict[str, dict[str, str]] = {
            config.sheet_name: {} for config in job.sheet_configs
        }
        # 每个 sheet 的结果级警告（如公式引用表头行上方单元格）
        sheet_extra_warnings: dict[str, list[str]] = {
            config.sheet_name: [] for config in job.sheet_configs
        }
        # 被追加过的 sheet：数据首行, 追加前真实末行, 追加后末行（用于图表范围延伸）
        extended_ranges: dict[str, tuple[int, int, int]] = {}

        total_units = max(1, len(job.input_files) * len(job.sheet_configs))
        done_units = 0

        try:
            for config in job.sheet_configs:
                unit_base = done_units / total_units
                done_units += len(job.input_files)

                def report(fraction: float, message: str) -> None:
                    percent = 5 + int(
                        min(1.0, unit_base + fraction * len(job.input_files) / total_units) * 90
                    )
                    progress.update(percent, message)

                if config.identical:
                    self._keep_identical_sheet(
                        workbook, config, job, source_rows[config.sheet_name], all_warnings
                    )
                    report(1.0, f"已保留 {config.sheet_name}（内容一致，不合并）")
                    continue
                sheet_plan = plans_by_sheet.get(config.sheet_name)
                if sheet_plan is None:
                    raise ValueError(f"找不到 sheet：{config.sheet_name}")
                try:
                    self._merge_sheet(
                        workbook,
                        config,
                        sheet_plan,
                        job,
                        source_rows[config.sheet_name],
                        skipped_duplicates[config.sheet_name],
                        fingerprints[config.sheet_name],
                        all_warnings,
                        sheet_extra_warnings[config.sheet_name],
                        extended_ranges,
                        report,
                    )
                except Exception as exc:
                    errors.append(f"{config.sheet_name}: {exc}")

            if extended_ranges:
                self._extend_chart_ranges(workbook, extended_ranges, all_warnings)

            ordered = [name for name in job.selected_sheets if name in workbook.sheetnames]
            if not ordered:
                raise ValueError("所选 sheet 在所有输入文件中都不存在")
            workbook._sheets = [workbook[name] for name in ordered]

            progress.update(97, "正在保存输出文件")
            job.output_file.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(job.output_file)
        finally:
            workbook.close()

        results: list[MergeSheetResult] = []
        for config in job.sheet_configs:
            sheet_plan = plans_by_sheet.get(config.sheet_name)
            sheet_source = source_rows[config.sheet_name]
            sheet_warnings: list[str] = list(sheet_extra_warnings[config.sheet_name])
            if sheet_plan is not None:
                for file_name, fields in sheet_plan.missing_fields.items():
                    if file_name in sheet_source:
                        sheet_warnings.append(
                            f"文件 {file_name} 缺少字段：{'、'.join(fields)}"
                        )
                for file_name, fields in sheet_plan.extra_fields.items():
                    if file_name in sheet_source:
                        sheet_warnings.append(
                            f"文件 {file_name} 多出字段：{'、'.join(fields)}，已追加到表头末尾"
                        )
            results.append(
                MergeSheetResult(
                    sheet_name=config.sheet_name,
                    merged_rows=sum(sheet_source.values()),
                    source_rows=sheet_source,
                    missing_fields=dict(sheet_plan.missing_fields) if sheet_plan else {},
                    extra_fields=dict(sheet_plan.extra_fields) if sheet_plan else {},
                    skipped_duplicates=dict(skipped_duplicates[config.sheet_name]),
                    warnings=sheet_warnings,
                )
            )
            all_warnings.extend(sheet_warnings)

        summary = MergeSummary(
            results=results,
            output_file=job.output_file,
            total_rows=sum(result.merged_rows for result in results),
            warnings=list(dict.fromkeys(all_warnings)),
            errors=errors,
        )
        progress.update(100, "合并完成")
        return summary

    def _keep_identical_sheet(
        self,
        workbook,
        config: MergeSheetConfig,
        job: MergeJob,
        source_rows: dict[str, int],
        warnings: list[str],
    ) -> None:
        """identical sheet：模板里有就原样保留；没有则从第一个包含的文件整体拷贝。

        其余文件的同名 sheet 完全不读取。
        """
        sheet_name = config.sheet_name
        if sheet_name in workbook.sheetnames:
            source_rows[job.input_files[0].name] = _count_data_rows(
                workbook[sheet_name], config.header_row
            )
            return
        for input_file in job.input_files[1:]:
            probe, _ = load_workbook_with_warnings(
                input_file, data_only=False, read_only=True
            )
            contains = sheet_name in probe.sheetnames
            probe.close()
            if not contains:
                continue
            ws = _copy_sheet_into_template(workbook, input_file, sheet_name)
            source_rows[input_file.name] = _count_data_rows(ws, config.header_row)
            return
        warnings.append(f"所有输入文件都不包含 sheet：{sheet_name}")

    def _merge_sheet(
        self,
        workbook,
        config: MergeSheetConfig,
        sheet_plan: MergeSheetPlan,
        job: MergeJob,
        source_rows: dict[str, int],
        skipped_duplicates: dict[str, str],
        fingerprints: dict[str, str],
        warnings: list[str],
        sheet_warnings: list[str],
        extended_ranges: dict[str, tuple[int, int, int]],
        report: Callable[[float, str], None],
    ) -> None:
        sheet_name = config.sheet_name
        ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        files = [
            path for path in job.input_files if path.name in sheet_plan.headers_by_file
        ]
        if not files:
            warnings.append(f"所有输入文件都不包含 sheet：{sheet_name}")
            return
        last_row_before_append: int | None = None
        total = len(files)
        for index, input_file in enumerate(files):
            if ws is None or input_file == job.input_files[0]:
                # 基准内容就位：模板自带则一个字节不动；模板缺该 sheet 时整体拷贝
                if ws is None:
                    ws = _copy_sheet_into_template(workbook, input_file, sheet_name)
                source_rows[input_file.name] = _count_data_rows(ws, config.header_row)
                if job.skip_duplicate_sheets:
                    # 指纹在截断前计算，与其他文件的流式指纹口径一致（都含幽灵行）
                    fingerprints[_sheet_fingerprint(ws)] = input_file.name
                _trim_ghost_rows(ws, config.header_row)
                self._prepare_union_columns(ws, config, sheet_plan, job, input_file)
                last_row_before_append = ws.max_row
                report((index + 1) / total, f"已就位 {input_file.name} / {sheet_name}")
                continue
            if job.skip_duplicate_sheets:
                fingerprint = _read_sheet_fingerprint(input_file, sheet_name)
                if fingerprint in fingerprints:
                    original = fingerprints[fingerprint]
                    skipped_duplicates[input_file.name] = original
                    warnings.append(
                        f"文件 {input_file.name} 的 sheet「{sheet_name}」"
                        f"与文件 {original} 内容完全相同，已跳过"
                    )
                    report((index + 1) / total, f"跳过重复 {input_file.name} / {sheet_name}")
                    continue
                fingerprints[fingerprint] = input_file.name
            added = self._append_file_rows(
                ws,
                input_file,
                config,
                sheet_plan,
                job,
                warnings,
                sheet_warnings,
                lambda fraction, message: report((index + fraction) / total, message),
            )
            source_rows[input_file.name] = added
            report((index + 1) / total, f"已合并 {input_file.name} / {sheet_name}")
        if (
            last_row_before_append is not None
            and ws is not None
            and ws.max_row > last_row_before_append
        ):
            extended_ranges[sheet_name] = (
                config.header_row + 1,
                last_row_before_append,
                ws.max_row,
            )

    def _prepare_union_columns(
        self,
        ws,
        config: MergeSheetConfig,
        sheet_plan: MergeSheetPlan,
        job: MergeJob,
        base_file: Path,
    ) -> None:
        """在模板表头行右侧补并集新增列和（可选的）来源列，并为模板自有数据行填来源。"""
        base_headers = sheet_plan.headers_by_file[base_file.name]
        base_ids = set(_identities(base_headers))
        extra_headers = [
            name
            for name, occ in _identities(sheet_plan.union_headers)
            if (name, occ) not in base_ids
        ]
        style_source = ws.cell(row=config.header_row, column=max(1, len(base_headers)))
        column = len(base_headers) + 1
        for header in extra_headers:
            cell = ws.cell(row=config.header_row, column=column, value=header)
            _copy_cell_style(style_source, cell)
            ws.column_dimensions[get_column_letter(column)].width = EXTRA_COLUMN_WIDTH
            column += 1
        if job.include_source_column:
            source_column = len(sheet_plan.union_headers) + 1
            cell = ws.cell(
                row=config.header_row, column=source_column, value=job.source_column_name
            )
            _copy_cell_style(style_source, cell)
            for row in ws.iter_rows(min_row=config.header_row + 1):
                if any(normalize_split_value(cell.value) is not None for cell in row):
                    ws.cell(row=row[0].row, column=source_column, value=base_file.stem)

    def _append_file_rows(
        self,
        ws,
        input_file: Path,
        config: MergeSheetConfig,
        sheet_plan: MergeSheetPlan,
        job: MergeJob,
        warnings: list[str],
        sheet_warnings: list[str],
        report: Callable[[float, str], None],
    ) -> int:
        """流式读取一个文件的数据行，按表头名映射追加到模板 sheet 末尾。"""
        sheet_name = config.sheet_name
        union_headers = sheet_plan.union_headers
        file_headers = sheet_plan.headers_by_file[input_file.name]
        mapping = _build_column_mapping(file_headers, union_headers)
        column_count = len(union_headers) + (1 if job.include_source_column else 0)
        layout = _read_sheet_layout(input_file, sheet_name, None)
        row_heights = layout["row_heights"]
        source_workbook, load_warnings = load_workbook_with_warnings(
            input_file, data_only=False, read_only=True
        )
        warnings.extend(load_warnings)
        added = 0
        consecutive_empty = 0
        next_row = ws.max_row + 1
        risky_references: dict[str, int] = {}  # 引用 -> 首次出现的源行号
        failed_translations = 0
        try:
            sheet = source_workbook[sheet_name]
            total_rows = (
                sheet.max_row - config.header_row if sheet.max_row is not None else None
            )
            for row in sheet.iter_rows(min_row=config.header_row + 1):
                is_empty = True
                pending: list[tuple[int, object, object]] = []  # (目标列, 值, 源单元格)
                for index, cell in enumerate(row):
                    if index >= len(mapping):
                        break
                    union_index = mapping[index]
                    if union_index is None:
                        continue
                    value = cell.value
                    if normalize_split_value(value) is not None:
                        is_empty = False
                    if value is None and not getattr(cell, "has_style", False):
                        continue
                    pending.append((union_index + 1, value, cell))
                if is_empty:
                    consecutive_empty += 1
                    if consecutive_empty >= EMPTY_ROW_STOP_THRESHOLD:
                        warnings.append(
                            f"文件 {input_file.name} 的 sheet「{sheet_name}」"
                            f"连续 {EMPTY_ROW_STOP_THRESHOLD} 行为空，已提前结束读取，"
                            f"后续空行将被忽略"
                        )
                        break
                    continue
                consecutive_empty = 0
                for target_column, value, source_cell in pending:
                    if isinstance(value, str) and value.startswith("="):
                        for ref in _find_risky_references(value, config.header_row):
                            risky_references.setdefault(ref, source_cell.row)
                        value, translated = _translate_formula(
                            value, source_cell.coordinate, target_column, next_row
                        )
                        if not translated:
                            failed_translations += 1
                    out_cell = ws.cell(row=next_row, column=target_column, value=value)
                    if getattr(source_cell, "has_style", False):
                        _copy_cell_style(source_cell, out_cell)
                if job.include_source_column:
                    ws.cell(row=next_row, column=column_count, value=input_file.stem)
                source_row_number = row[0].row if row else None
                if source_row_number in row_heights:
                    ws.row_dimensions[next_row].height = row_heights[source_row_number]
                next_row += 1
                added += 1
                if added % 2000 == 0:
                    fraction = added / total_rows if total_rows else 0.0
                    message = (
                        f"正在合并 {input_file.name} / {sheet_name} ({added}/{total_rows})"
                        if total_rows
                        else f"正在合并 {input_file.name} / {sheet_name}"
                    )
                    report(fraction, message)
        finally:
            source_workbook.close()
        for ref, row_number in risky_references.items():
            parts = re.fullmatch(r"([A-Z]+)([0-9]+)", ref)
            absolute = f"${parts.group(1)}${parts.group(2)}" if parts else ref
            sheet_warnings.append(
                f"文件 {input_file.name} 的 sheet「{sheet_name}」第 {row_number} 行公式"
                f"引用了表头行上方的单元格 {ref}，平移后可能指向错误位置，"
                f"建议改用绝对引用（{absolute}）"
            )
        if failed_translations:
            sheet_warnings.append(
                f"文件 {input_file.name} 的 sheet「{sheet_name}」有 "
                f"{failed_translations} 个公式平移失败，已保留原公式，请检查"
            )
        return added

    def _extend_chart_ranges(
        self,
        workbook,
        extended_ranges: dict[str, tuple[int, int, int]],
        warnings: list[str],
    ) -> None:
        """把图表中指向"被追加 sheet 的追加前末行"的数据区域延伸到追加后末行。

        只处理末行恰好等于追加前末行的完整区域，其余一律不动（不瞎猜部分区间）。
        任何异常都不中断合并，记警告继续。
        """
        for ws in workbook.worksheets:
            for chart in getattr(ws, "_charts", None) or []:
                try:
                    self._extend_chart(chart, ws.title, extended_ranges)
                except Exception as exc:
                    warnings.append(
                        f"延伸图表数据范围失败（{ws.title}）：{exc}，已保留原范围"
                    )

    def _extend_chart(
        self, chart, current_sheet: str, extended_ranges: dict[str, tuple[int, int, int]]
    ) -> None:
        for series in getattr(chart, "series", None) or []:
            for ref_holder in _series_ref_holders(series):
                formula = getattr(ref_holder, "f", None)
                updated = _extend_ref_formula(formula, current_sheet, extended_ranges)
                if updated is not None:
                    ref_holder.f = updated


def _series_ref_holders(series) -> list:
    """取 series 上的数据区域引用对象（值/分类/散点 XY），不含单单元格的标题引用。"""
    holders = []
    val = getattr(series, "val", None)
    ref = getattr(val, "numRef", None)
    if ref is not None:
        holders.append(ref)
    cat = getattr(series, "cat", None)
    for attr in ("strRef", "numRef"):
        ref = getattr(cat, attr, None)
        if ref is not None:
            holders.append(ref)
    for attr in ("xVal", "yVal"):
        axis = getattr(series, attr, None)
        ref = getattr(axis, "numRef", None)
        if ref is not None:
            holders.append(ref)
    return holders


def _extend_ref_formula(
    formula, current_sheet: str, extended_ranges: dict[str, tuple[int, int, int]]
) -> str | None:
    """若引用指向被追加 sheet 且覆盖数据区，返回延伸后的引用，否则 None。

    延伸规则：区域起始行 == 数据首行（表头行+1）且末行 >= 追加前真实末行
    （覆盖图表范围大于真实数据的场景，如拆分保留的原表大区间）。
    """
    if not isinstance(formula, str) or not formula:
        return None
    if "!" in formula:
        prefix, range_part = formula.rsplit("!", 1)
        sheet_name = prefix.strip("'")
    else:
        prefix, range_part, sheet_name = None, formula, current_sheet
    if sheet_name not in extended_ranges:
        return None
    data_start, before, after = extended_ranges[sheet_name]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(
            range_part.replace("$", "")
        )
    except (ValueError, TypeError):
        return None
    if max_row is None or (min_col == max_col and min_row == max_row):
        return None  # 整列引用或单单元格引用不处理
    if min_row != data_start or max_row < before:
        return None  # 不从数据首行开始、或不覆盖到数据末行的不瞎猜
    new_range = (
        f"${get_column_letter(min_col)}${min_row}"
        f":${get_column_letter(max_col)}${after}"
    )
    return f"{prefix}!{new_range}" if prefix is not None else new_range


def _translate_formula(
    formula: str, origin: str, target_column: int, target_row: int
) -> tuple[str, bool]:
    """把公式从源坐标平移到目标坐标，返回（公式, 是否成功）；失败时保留原公式。"""
    try:
        return (
            Translator(formula, origin=origin).translate_formula(
                f"{get_column_letter(target_column)}{target_row}"
            ),
            True,
        )
    except Exception:
        return formula, False


def _build_column_mapping(
    file_headers: list[str | None], union_headers: list[str]
) -> list[int | None]:
    """按 (名字, 同名字段出现序号) 把文件列映射到并集列序，同名字段不塌缩。"""
    union_positions: dict[str, list[int]] = {}
    for index, header in enumerate(union_headers):
        union_positions.setdefault(header, []).append(index)
    file_counts: dict[str, int] = {}
    mapping: list[int | None] = []
    for header in file_headers:
        if header is None:
            mapping.append(None)
            continue
        occurrence = file_counts.get(header, 0)
        file_counts[header] = occurrence + 1
        positions = union_positions.get(header, [])
        mapping.append(positions[occurrence] if occurrence < len(positions) else None)
    return mapping


def _trim_ghost_rows(worksheet, header_row: int) -> None:
    """截掉数据区尾部的幽灵行（拆分产物里被删数据留下的全空带样式行）。

    整行值全 None 才算幽灵行；公式字符串算非空，不会被误删。
    """
    real_end = _real_last_data_row(worksheet, header_row)
    if worksheet.max_row > real_end:
        worksheet.delete_rows(real_end + 1, worksheet.max_row - real_end)


def _real_last_data_row(worksheet, header_row: int) -> int:
    last_row = header_row
    cells = getattr(worksheet, "_cells", None)
    if isinstance(cells, dict):
        for (row_index, _column), cell in cells.items():
            if row_index > header_row and cell.value is not None:
                last_row = max(last_row, row_index)
        return last_row
    for row in worksheet.iter_rows(min_row=header_row + 1):
        if any(cell.value is not None for cell in row):
            last_row = row[0].row
    return last_row


def _find_risky_references(formula: str, header_row: int) -> list[str]:
    """找出公式中"同 sheet、行号 <= 表头行的相对引用"（平移后可能指错位置）。

    只处理常见 A1 形态；跨 sheet 引用（带 !）、绝对引用（带 $）、函数名、
    区域右半部分都不纳入（宁缺毋滥）。
    """
    refs: list[str] = []
    for match in _CELL_REF_PATTERN.finditer(formula):
        column_abs, _column, row_abs, row_digits = match.groups()
        if column_abs or row_abs:
            continue  # 绝对引用不警告
        if int(row_digits) <= header_row:
            ref = match.group(0)
            if ref not in refs:
                refs.append(ref)
    return refs


def _count_data_rows(worksheet, header_row: int) -> int:
    count = 0
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if any(normalize_split_value(value) is not None for value in row):
            count += 1
    return count


def _copy_sheet_into_template(workbook, input_file: Path, sheet_name: str):
    """openpyxl 不能跨工作簿复制 sheet：逐格复制值+样式，再补列宽/行高/冻结/合并。"""
    layout = _read_sheet_layout(input_file, sheet_name, None)
    source_workbook, _ = load_workbook_with_warnings(
        input_file, data_only=False, read_only=True
    )
    try:
        source = source_workbook[sheet_name]
        target = workbook.create_sheet(sheet_name)
        for row in source.iter_rows():
            for cell in row:
                if cell.value is None and not getattr(cell, "has_style", False):
                    continue
                out_cell = target.cell(row=cell.row, column=cell.column, value=cell.value)
                if getattr(cell, "has_style", False):
                    _copy_cell_style(cell, out_cell)
    finally:
        source_workbook.close()
    for letter, width in layout["widths"].items():
        target.column_dimensions[letter].width = width
    for row_number, height in layout["row_heights"].items():
        target.row_dimensions[row_number].height = height
    if layout["freeze_panes"]:
        target.freeze_panes = layout["freeze_panes"]
    for ref in layout["merges"]:
        target.merge_cells(ref)
    return target


def _copy_cell_style(source_cell, target_cell) -> None:
    try:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.border = copy(source_cell.border)
        target_cell.protection = copy(source_cell.protection)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
    except Exception:
        pass  # 复制不了的样式项直接放弃，不影响合并结果


def _sheet_fingerprint(worksheet) -> str:
    digest = hashlib.sha256()
    for row in worksheet.iter_rows(values_only=True):
        digest.update(b"\x1e")  # 行分隔
        for value in row:
            digest.update(_fingerprint_token(value))
    return digest.hexdigest()


def _read_sheet_fingerprint(path: Path, sheet_name: str) -> str:
    """流式读取一个文件的指定 sheet 并计算内容指纹。

    基于读取到的原始行值，列顺序不同不算重复（保守语义，仍按表头名对齐合并）。
    """
    workbook, _ = load_workbook_with_warnings(path, data_only=False, read_only=True)
    try:
        return _sheet_fingerprint(workbook[sheet_name])
    finally:
        workbook.close()


def _fingerprint_token(value) -> bytes:
    if value is None:
        return b"\x00"
    # 带上类型名，避免数字 1 与字符串 "1" 被当成相同内容
    return b"\x01" + f"{type(value).__name__}:{value}".encode("utf-8", "replace") + b"\x1f"


def _sheet_archive_entry(archive: ZipFile, sheet_name: str) -> str | None:
    """在 xlsx 压缩包里定位指定 sheet 的 XML 条目。"""
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    workbook_xml = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    rel_id = None
    for sheet_node in workbook_xml.iter(f"{main_ns}sheet"):
        if sheet_node.get("name") == sheet_name:
            rel_id = sheet_node.get(f"{rel_ns}id")
            break
    if rel_id is None:
        return None
    rels_xml = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for rel_node in rels_xml:
        if rel_node.get("Id") == rel_id:
            target = rel_node.get("Target")
            if target:
                return target.lstrip("/") if target.startswith("/") else f"xl/{target}"
    return None


def _read_sheet_layout(path: Path, sheet_name: str, max_height_row: int | None) -> dict:
    """直接从 xlsx 压缩包解析 sheet 布局（read_only 模式不提供这些属性）。

    解析列宽、冻结窗格、行高（max_height_row 为 None 时全部行，否则只到该
    行）、合并单元格。解析失败时返回空布局，对应项直接放弃，不影响合并结果。
    """
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    layout: dict = {"widths": {}, "freeze_panes": None, "row_heights": {}, "merges": []}
    try:
        with ZipFile(path, "r") as archive:
            sheet_entry = _sheet_archive_entry(archive, sheet_name)
            if sheet_entry is None:
                return layout
            with archive.open(sheet_entry) as stream:
                for event, node in ElementTree.iterparse(stream, events=("start", "end")):
                    tag = node.tag
                    if event == "start" and tag == f"{main_ns}pane":
                        if node.get("state") == "frozen":
                            layout["freeze_panes"] = node.get("topLeftCell")
                    elif event == "end" and tag == f"{main_ns}col":
                        width = float(node.get("width", 0))
                        if width:
                            for column in range(
                                int(node.get("min")), int(node.get("max")) + 1
                            ):
                                layout["widths"][get_column_letter(column)] = width
                    elif event == "start" and tag == f"{main_ns}row":
                        height = node.get("ht")
                        row_number = int(node.get("r", 0))
                        if height and (
                            max_height_row is None or row_number <= max_height_row
                        ):
                            layout["row_heights"][row_number] = float(height)
                    elif event == "end" and tag == f"{main_ns}mergeCell":
                        ref = node.get("ref")
                        if ref:
                            layout["merges"].append(ref)
    except Exception:
        pass
    return layout


class _ProgressReporter:
    def __init__(self, callback: ProgressCallback | None) -> None:
        self.callback = callback
        self.last_percent = -1

    def update(self, percent: int, message: str) -> None:
        normalized = max(self.last_percent, min(100, max(0, percent)))
        if self.callback is not None and (
            normalized != self.last_percent or normalized in (0, 100)
        ):
            self.callback(normalized, message)
        self.last_percent = normalized
