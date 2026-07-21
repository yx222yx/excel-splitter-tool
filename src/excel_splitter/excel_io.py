from __future__ import annotations

from pathlib import Path
import warnings

from openpyxl import load_workbook

from .values import normalize_split_value


def load_workbook_with_warnings(path: Path, **kwargs):
    with warnings.catch_warnings(record=True) as captured:
        warnings.simplefilter("always")
        workbook = load_workbook(path, **kwargs)
    messages = list(dict.fromkeys(str(item.message) for item in captured))
    return workbook, messages


def detect_blocks(sheet, header_row: int) -> list[tuple[int, int, int]]:
    """把表头行之下的数据区按「整行全空」切成若干块（拆分/合并共用）。

    块 1 = 主表（表头行 + 其下连续非空行）；每个空行之后若还有非空行则开始
    新块，块首行视为块表头。返回 [(块表头行, 块数据首行, 块数据末行), ...]，
    单表区 sheet 只有块 1。检测用读到的值（调用方保证 data_only 口径一致）。
    """
    max_row = sheet.max_row or header_row
    segments: list[tuple[int, int]] = []
    segment_start: int | None = None
    last_non_empty = header_row
    consecutive_empty = 0
    EMPTY_GAP_STOP = 10000  # 连续空行上限，超出视为后续无数据
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, max_row=max_row, values_only=True),
        start=header_row + 1,
    ):
        if all(normalize_split_value(value) is None for value in row):
            consecutive_empty += 1
            if consecutive_empty >= EMPTY_GAP_STOP:
                break
            if segment_start is not None:
                segments.append((segment_start, last_non_empty))
                segment_start = None
        else:
            consecutive_empty = 0
            if segment_start is None:
                segment_start = row_index
            last_non_empty = row_index
    if segment_start is not None:
        segments.append((segment_start, last_non_empty))
    if not segments:
        return [(header_row, header_row + 1, header_row)]
    blocks = [(header_row, header_row + 1, segments[0][1])]
    for start, end in segments[1:]:
        blocks.append((start, start + 1, end))
    return blocks

