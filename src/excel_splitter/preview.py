from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter

from .excel_io import load_workbook_with_warnings
from .models import MAX_HEADER_ROW


@dataclass(frozen=True, slots=True)
class ColumnOption:
    index: int
    letter: str
    header: str

    @property
    def label(self) -> str:
        return f"{self.letter} - {self.header}"


@dataclass(frozen=True, slots=True)
class SheetPreview:
    sheet_name: str
    rows: tuple[tuple[Any, ...], ...]
    suggested_header_row: int
    start_row: int
    end_row: int
    total_rows: int
    has_more: bool
    warnings: tuple[str, ...] = ()

    def columns_for_header(self, header_row: int) -> list[ColumnOption]:
        row_offset = header_row - self.start_row
        if row_offset < 0 or row_offset >= len(self.rows):
            raise ValueError("表头行不在预览范围内")
        row = self.rows[row_offset]
        last_nonempty = max(
            (index for index, value in enumerate(row, start=1) if value is not None),
            default=0,
        )
        return [
            ColumnOption(
                index=index,
                letter=get_column_letter(index),
                header=str(row[index - 1]).strip() if row[index - 1] is not None else "(空列)",
            )
            for index in range(1, last_nonempty + 1)
        ]


def list_sheet_names(input_file: Path) -> list[str]:
    workbook, _ = load_workbook_with_warnings(input_file, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def preview_sheet(
    input_file: Path,
    sheet_name: str,
    *,
    start_row: int = 1,
    max_rows: int = 100,
) -> SheetPreview:
    if start_row < 1:
        raise ValueError("预览起始行必须大于等于 1")
    if max_rows < 1 or max_rows > 500:
        raise ValueError("单次预览行数必须在 1 到 500 之间")
    workbook, warning_messages = load_workbook_with_warnings(
        input_file, data_only=True, read_only=True
    )
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"找不到 sheet：{sheet_name}")
        sheet = workbook[sheet_name]
        total_rows = sheet.max_row
        end_row = min(start_row + max_rows - 1, total_rows)
        rows = tuple(
            tuple(cell.value for cell in row)
            for row in sheet.iter_rows(min_row=start_row, max_row=end_row)
        )
        header_rows = tuple(
            tuple(cell.value for cell in row)
            for row in sheet.iter_rows(
                min_row=1, max_row=min(MAX_HEADER_ROW, total_rows)
            )
        )
    finally:
        workbook.close()
    return SheetPreview(
        sheet_name=sheet_name,
        rows=rows,
        suggested_header_row=suggest_header_row(header_rows),
        start_row=start_row,
        end_row=end_row,
        total_rows=total_rows,
        has_more=end_row < total_rows,
        warnings=tuple(warning_messages),
    )


def suggest_header_row(rows: tuple[tuple[Any, ...], ...]) -> int:
    candidates = rows[:MAX_HEADER_ROW]
    if not candidates:
        return 1
    width = max((len(row) for row in candidates), default=1)
    scored: list[tuple[float, int]] = []
    for row_index, row in enumerate(candidates, start=1):
        values = [value for value in row if value is not None and str(value).strip()]
        if not values:
            scored.append((-100.0, row_index))
            continue
        strings = sum(isinstance(value, str) for value in values)
        score = (len(values) / width) * 2
        if strings / len(values) >= 0.7:
            score += 2
        if len({str(value) for value in values}) == len(values):
            score += 1
        next_row = rows[row_index] if row_index < len(rows) else ()
        next_values = [
            value for value in next_row if value is not None and str(value).strip()
        ]
        if len(next_values) >= 2:
            score += 1
        if any(isinstance(value, (int, float)) for value in next_values):
            score += 1
        if len(values) == 1 and len(next_values) >= 2:
            score -= 3
        scored.append((score, row_index))
    return max(scored, key=lambda item: (item[0], -item[1]))[1]
